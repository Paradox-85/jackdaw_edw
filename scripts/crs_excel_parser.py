"""
EIS Excel Parser — CRS Comments Extractor
Reads DOC_COMMENT_* (Group 1) and JDAW-*_N.xlsx (Group 2) files,
merges them into a fixed-schema master DataFrame and saves to CSV + Excel sample.

Requirements:
    pip install pandas openpyxl python-calamine
"""

import re
import logging
import threading
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _scalar(val):
    """Extract a scalar from a value that may be a pd.Series."""
    if isinstance(val, pd.Series):
        return val.iloc[0] if not val.empty else None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def _norm_sheet(name: str) -> str:
    """Normalise sheet name: strip, lowercase, spaces → underscores."""
    return name.strip().lower().replace(" ", "_")


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SOURCE_DIR = Path(
    r"C:\Users\ADZV\OneDrive - Ramboll\Ramboll_Jackdaw - Admin Team\EIS\Export for Shell"
)
OUTPUT_PATH = Path(
    r"C:\Users\ADZV\OneDrive - Ramboll\Ramboll_Jackdaw - Admin Team"
    r"\Information Management\EDW\master_crs_register.csv"
)
OUTPUT_SAMPLE_PATH = OUTPUT_PATH.with_name("master_crs_register_sample.xlsx")

MAX_WORKERS       = 6
EXCEL_SAMPLE_ROWS = 50_000

OUTPUT_COLS = [
    "DOC_NUMBER",
    "REVISION",
    "RETURN_CODE",
    "TRANSMITTAL_NUMBER",
    "TRANSMITTAL_DATE",
    "TAG_NAME",
    "PROPERTY_NAME",
    "GROUP_COMMENT",
    "RESPONSE",
    "SOURCE_FILE",
    "DETAIL_FILE",
    "DETAIL_SHEET",
    "COMMENT",
]

SKIP_SHEETS = {_norm_sheet("Comment Sheet")}

# FIX 4: correct raw strings — single backslash
MAIN_PATTERN = re.compile(
    r"^DOC_COMMENT_(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d{2})_[A-Z]{3}\.xlsx$"
)
DETAIL_PATTERN = re.compile(
    r"^(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d{2})(?:_\d+|_Review_Comments)\.xlsx$"
)

COMMENT_COL_KEYWORDS  = ("remark", "adura", "issue", "comment")
PROPERTY_COL_KEYWORDS = ("equipment property name", "tag property name", "property name")

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------
def discover_files(
    root: Path,
) -> tuple[dict[str, Path], dict[str, list[Path]]]:
    """Recursively scan root, skip _templates at any depth."""
    main_files: dict[str, Path] = {}
    detail_files: dict[str, list[Path]] = {}

    for path in root.rglob("*.xlsx"):
        if "_templates" in path.parts:
            continue

        name = path.name
        m = MAIN_PATTERN.match(name)
        if m:
            key = m.group(1)
            if key in main_files:
                log.warning("Duplicate main file key %s — keeping: %s", key, main_files[key])
            else:
                main_files[key] = path
            continue

        d = DETAIL_PATTERN.match(name)
        if d:
            detail_files.setdefault(d.group(1), []).append(path)

    log.info("Found %d main file(s), %d detail key(s).", len(main_files), len(detail_files))
    return main_files, detail_files


# ---------------------------------------------------------------------------
# Merged-cell resolution (openpyxl)
# ---------------------------------------------------------------------------
def _expand_merged_cells(ws) -> dict[tuple[int, int], object]:
    cell_map: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for cell in row:
            cell_map[(cell.row, cell.column)] = cell.value

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = cell_map.get((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_map[(r, c)] = top_left_val

    return cell_map


def _build_two_row_header(
    cell_map: dict, row1: int, row2: int, max_col: int
) -> list[str]:
    """Combine two header rows into UPPER_SNAKE column names. Deduplicates."""
    headers = []
    for col in range(1, max_col + 1):
        v1 = str(cell_map.get((row1, col), "") or "").strip()
        v2 = str(cell_map.get((row2, col), "") or "").strip()

        if v1 and v2 and v1 != v2:
            name = f"{v1}_{v2}"
        elif v1:
            name = v1
        elif v2:
            name = v2
        else:
            name = f"COL_{col}"

        # FIX: correct character class — single backslash before s
        name = re.sub(r"[\s/\\().,-]+", "_", name).upper().strip("_")
        headers.append(name)

    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


# ---------------------------------------------------------------------------
# Detect the COMMENT column in a detail sheet via merged-cell analysis
# ---------------------------------------------------------------------------
def _find_comment_column(ws, data_start_row: int) -> int | None:
    """
    Detect a vertically-merged single-column comment cell.
    - Single-column merge only (min_col == max_col)
    - Must start exactly at data_start_row
    - Must span at least 2 rows
    - If multiple candidates: pick the one with the largest span.
    Returns 0-based column index, or None.
    """
    best_span: int = 1
    best_col:  int | None = None

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row_range = merged_range.bounds

        if min_col != max_col:
            continue
        if min_row != data_start_row:
            continue

        span = max_row_range - min_row + 1
        if span > best_span:
            best_span = span
            best_col  = min_col - 1

    return best_col


# ---------------------------------------------------------------------------
# Main-file parser
# ---------------------------------------------------------------------------
def parse_main_file(
    path: Path,
) -> tuple[dict | None, pd.DataFrame | None]:
    """
    Extract header metadata and comment table from a DOC_COMMENT_* file.
    Header rows 1–5 (openpyxl), data rows 8+ (calamine).
    """
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.active
    except Exception as exc:
        log.error("Cannot open workbook %s: %s", path.name, exc)
        return None, None

    cell_map = _expand_merged_cells(ws)
    max_col  = ws.max_column or 0
    wb.close()

    def g(row: int, col: int):
        val = cell_map.get((row, col))
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        return val

    metadata = {
        "DOC_NUMBER":         g(4, 3),
        "REVISION":           g(3, 6),
        "RETURN_CODE":        g(4, 6),
        "TRANSMITTAL_NUMBER": g(5, 3),
        "TRANSMITTAL_DATE":   g(5, 6),
        "SOURCE_FILE":        path.name,
    }

    rc = str(metadata["RETURN_CODE"]).strip().split(".")[0]
    if rc == "1":
        log.info("Skipping %s — Return Code 1.", path.name)
        return None, None

    if max_col < 3:
        log.warning("Too few columns in %s — skipping.", path.name)
        return metadata, None

    headers = _build_two_row_header(cell_map, row1=6, row2=7, max_col=max_col)

    try:
        df_raw = pd.read_excel(
            path, header=None, skiprows=7,
            engine="calamine", dtype=str, na_filter=False,
        )
    except Exception as exc:
        log.error("Cannot read data rows of %s: %s", path.name, exc)
        return metadata, None

    if df_raw.empty:
        log.warning("No data rows in %s.", path.name)
        return metadata, None

    actual_cols = df_raw.shape[1]
    if len(headers) < actual_cols:
        headers += [f"COL_{i}" for i in range(len(headers), actual_cols)]
    else:
        headers = headers[:actual_cols]
    df_raw.columns = headers

    col_c_name = headers[2] if len(headers) > 2 else None
    col_f_name = headers[5] if len(headers) > 5 else None

    if col_c_name is None:
        log.warning("Column C missing in %s.", path.name)
        return metadata, None

    df_comments = pd.DataFrame()
    df_comments["GROUP_COMMENT"] = df_raw[col_c_name]
    df_comments["RESPONSE"]      = df_raw[col_f_name] if col_f_name else ""

    col_a_name   = headers[0] if headers else None
    mask_a_empty = (
        df_raw[col_a_name].str.strip() == ""
        if col_a_name else pd.Series([True] * len(df_raw))
    )
    mask_c_empty = df_comments["GROUP_COMMENT"].str.strip() == ""

    df_comments = df_comments[~(mask_a_empty & mask_c_empty)].copy()
    df_comments = df_comments[
        df_comments["GROUP_COMMENT"].str.strip() != ""
    ].reset_index(drop=True)
    df_comments["RESPONSE"] = df_comments["RESPONSE"].replace("", None)

    return metadata, df_comments


# ---------------------------------------------------------------------------
# Detail-file loader — thread-safe, two-phase
# ---------------------------------------------------------------------------

# SheetData: (DataFrame, merged_comment_value, fallback_comment_col_name)
SheetData = tuple[pd.DataFrame, str | None, str | None]


def _load_detail_file_impl(path: Path) -> dict[str, SheetData]:
    """
    Phase 1 — openpyxl: detect vertically-merged comment column per sheet.
    Phase 2 — calamine: read data rows fast, drop merged comment col from df.
    Returns {sheet_key: (df, merged_comment_value, fallback_comment_col)}.
    """
    result: dict[str, SheetData] = {}

    # Phase 1: openpyxl
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        log.error("Cannot open detail file %s: %s", path.name, exc)
        return result

    sheet_names    = list(wb.sheetnames)
    DATA_START_ROW = 2
    sheet_meta: dict[str, tuple[int | None, str | None]] = {}

    for sheet in sheet_names:
        sheet_key = _norm_sheet(sheet)
        if sheet_key in SKIP_SHEETS:
            log.info("  Skipping sheet '%s' in %s", sheet, path.name)
            continue

        ws = wb[sheet]
        comment_col_idx = _find_comment_column(ws, data_start_row=DATA_START_ROW)

        comment_value: str | None = None
        if comment_col_idx is not None:
            raw = ws.cell(row=DATA_START_ROW, column=comment_col_idx + 1).value
            comment_value = str(raw).strip() if raw is not None else None
            comment_value = comment_value or None
            log.info("  Sheet '%s': merged comment col=%d val=%r",
                     sheet, comment_col_idx, str(comment_value)[:60])
        else:
            log.info("  Sheet '%s': no merged comment column.", sheet)

        sheet_meta[sheet_key] = (comment_col_idx, comment_value)

    wb.close()  # closed exactly once

    # Phase 2: calamine
    try:
        xl = pd.ExcelFile(path, engine="calamine")
    except Exception as exc:
        log.error("Cannot open detail file with calamine %s: %s", path.name, exc)
        return result

    for sheet in sheet_names:
        sheet_key = _norm_sheet(sheet)
        if sheet_key not in sheet_meta:
            continue

        comment_col_idx, comment_value = sheet_meta[sheet_key]

        try:
            df = pd.read_excel(
                xl, sheet_name=sheet, engine="calamine",
                dtype=str, na_filter=False,
            )
        except Exception as exc:
            log.warning("Cannot read sheet '%s' in %s: %s", sheet, path.name, exc)
            continue

        df = df[~df.apply(lambda r: all(v == "" for v in r), axis=1)].copy()
        if df.empty:
            log.warning("  Sheet '%s' in %s empty after filtering.", sheet, path.name)
            continue

        # Drop merged comment column — value already captured as scalar
        if comment_col_idx is not None and comment_col_idx < len(df.columns):
            df = df.drop(columns=df.columns[comment_col_idx])

        # FIX 2 & 3: fallback per-row comment column when no merge detected
        fallback_comment_col: str | None = None
        if comment_col_idx is None:
            for col in df.columns:
                if any(kw in col.lower() for kw in COMMENT_COL_KEYWORDS):
                    fallback_comment_col = col
                    log.info("  Sheet '%s': fallback comment col=%r", sheet, col)
                    break

        # KEY FIX: for per-row comment columns, filter out rows where comment is empty.
        # These rows have no finding/error — no reason to include them in output.
        # Do NOT filter for merged comments (comment_value is not None) —
        # the merge applies to all rows by definition.
        if fallback_comment_col is not None:
            before = len(df)
            df = df[df[fallback_comment_col].str.strip() != ""].copy()
            df = df.reset_index(drop=True)
            log.info(
                "  Sheet '%s': filtered by fallback col '%s' — %d → %d rows",
                sheet, fallback_comment_col, before, len(df),
            )
            if df.empty:
                log.warning("  Sheet '%s': no rows with non-empty comment, skipping.", sheet)
                continue

        # FIX 1: store 3-tuple consistently
        result[sheet_key] = (df, comment_value, fallback_comment_col)
        log.info("  Sheet '%s' loaded: %d row(s)", sheet, len(df))

    return result


_detail_file_cache: dict[Path, dict] = {}
_cache_lock = threading.Lock()


def _load_detail_file(path: Path) -> dict[str, SheetData]:
    """Thread-safe cache wrapper — double-checked locking."""
    if path in _detail_file_cache:
        return _detail_file_cache[path]
    with _cache_lock:
        if path not in _detail_file_cache:
            _detail_file_cache[path] = _load_detail_file_impl(path)
    return _detail_file_cache[path]


# FIX 2: unpack 3-tuple correctly
def find_matching_sheet(
    comment_text: str,
    sheets: dict[str, SheetData],
) -> tuple[str | None, pd.DataFrame | None, str | None, str | None]:
    """
    Returns (sheet_key, df, merged_comment_value, fallback_comment_col).
    Normalises comment_text same way as sheet keys (spaces → underscores).
    """
    text_norm = comment_text.lower().replace(" ", "_")
    for sheet_key, (df, comment_val, fallback_col) in sheets.items():  # unpack 3-tuple
        if sheet_key in text_norm:
            return sheet_key, df, comment_val, fallback_col
    return None, None, None, None


# ---------------------------------------------------------------------------
# Per-key processing unit — runs in thread
# ---------------------------------------------------------------------------
def process_key(
    key: str,
    main_path: Path,
    related_paths: tuple[Path, ...],
) -> list[dict]:
    """Returns flat list of records with exactly OUTPUT_COLS fields."""
    records: list[dict] = []

    metadata, df_comments = parse_main_file(main_path)
    if metadata is None or df_comments is None:
        return records

    all_detail_sheets: dict[Path, dict] = {
        p: _load_detail_file(p) for p in related_paths
    }

    for _, row in df_comments.iterrows():
        comment_text  = str(row["GROUP_COMMENT"]).strip()
        response_text = _scalar(row.get("RESPONSE"))
        found_detail  = False

        for detail_path, sheets in all_detail_sheets.items():
            # FIX 3: unpack all 4 return values
            sheet_name, df_sheet, comment_val, fallback_col = find_matching_sheet(
                comment_text, sheets
            )
            if df_sheet is None:
                continue

            found_detail = True

            # TAG_NAME: column containing "tag" but not "property"
            tag_col = next(
                (c for c in df_sheet.columns
                 if "tag" in c.lower() and "property" not in c.lower()), None
            )

            # FIX 5: PROPERTY_NAME column detection
            prop_col = next(
                (c for c in df_sheet.columns
                 if any(kw in c.lower() for kw in PROPERTY_COL_KEYWORDS)), None
            )

            for _, d_row in df_sheet.iterrows():
                tag_name  = _scalar(d_row[tag_col])  if tag_col  else None
                prop_name = _scalar(d_row[prop_col]) if prop_col else None
                if not prop_name or str(prop_name).strip() == "":
                    prop_name = "Not Applicable"

                # COMMENT: merged scalar takes priority; fallback = per-row value
                row_comment = comment_val
                if row_comment is None and fallback_col and fallback_col in d_row.index:
                    row_comment = _scalar(d_row[fallback_col]) or None

                records.append({
                    "DOC_NUMBER":         metadata.get("DOC_NUMBER"),
                    "REVISION":           metadata.get("REVISION"),
                    "RETURN_CODE":        metadata.get("RETURN_CODE"),
                    "TRANSMITTAL_NUMBER": metadata.get("TRANSMITTAL_NUMBER"),
                    "TRANSMITTAL_DATE":   metadata.get("TRANSMITTAL_DATE"),
                    "TAG_NAME":           tag_name,
                    "PROPERTY_NAME":      prop_name,   # FIX 5
                    "GROUP_COMMENT":      comment_text,
                    "RESPONSE":           response_text,
                    "SOURCE_FILE":        metadata.get("SOURCE_FILE"),
                    "DETAIL_FILE":        detail_path.name,
                    "DETAIL_SHEET":       sheet_name,
                    "COMMENT":            row_comment,  # FIX 3
                })

            break  # one comment → one matching sheet

        if not found_detail:
            records.append({
                "DOC_NUMBER":         metadata.get("DOC_NUMBER"),
                "REVISION":           metadata.get("REVISION"),
                "RETURN_CODE":        metadata.get("RETURN_CODE"),
                "TRANSMITTAL_NUMBER": metadata.get("TRANSMITTAL_NUMBER"),
                "TRANSMITTAL_DATE":   metadata.get("TRANSMITTAL_DATE"),
                "TAG_NAME":           None,
                "PROPERTY_NAME":      "Not Applicable",
                "GROUP_COMMENT":      comment_text,
                "RESPONSE":           response_text,
                "SOURCE_FILE":        metadata.get("SOURCE_FILE"),
                "DETAIL_FILE":        None,
                "DETAIL_SHEET":       None,
                "COMMENT":            "No matching detail sheet found",
            })

    return records


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------
def process() -> pd.DataFrame:
    main_files, detail_files = discover_files(SOURCE_DIR)

    work_items = [
        (key, path, tuple(detail_files.get(key, [])))
        for key, path in main_files.items()
    ]

    log.info(
        "Starting parallel processing — %d worker(s), %d job(s).",
        MAX_WORKERS, len(work_items),
    )

    all_records: list[dict] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(process_key, key, path, details): key
            for key, path, details in work_items
        }
        for future in as_completed(futures):
            key = futures[future]
            try:
                records = future.result()
                all_records.extend(records)
                log.info("  ✓ %s — %d record(s)", key, len(records))
            except Exception as exc:
                log.error("  ✗ %s failed: %s", key, exc)

    result_df = pd.DataFrame(all_records, columns=OUTPUT_COLS)

    if not result_df.empty:
        top = (
            result_df.groupby("SOURCE_FILE")["GROUP_COMMENT"]
            .count()
            .sort_values(ascending=False)
            .head(10)
        )
        log.info("Top files by record count:\n%s", top.to_string())

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    result_df.to_csv(
        OUTPUT_PATH, index=False, encoding="utf-8-sig",
        sep=",", date_format="%Y-%m-%d",
    )
    log.info("CSV saved (%d rows × %d cols) → %s",
             len(result_df), len(result_df.columns), OUTPUT_PATH)

    sample_df = result_df.head(EXCEL_SAMPLE_ROWS)
    sample_df.to_excel(OUTPUT_SAMPLE_PATH, index=False, engine="openpyxl")
    log.info("Excel sample saved (%d rows × %d cols) → %s",
             len(sample_df), len(sample_df.columns), OUTPUT_SAMPLE_PATH)

    return result_df


if __name__ == "__main__":
    process()