"""
import_crs_data.py

One-shot script: parse CRS (Comment Review Sheet) Excel files and load into
audit_core.crs_comment (PostgreSQL).

Run manually:
    python scripts/import_crs_data.py
    python scripts/import_crs_data.py --debug            # latest revision only
    python scripts/import_crs_data.py --debug-rev A28    # specific revision

Parsing logic ported from scripts/crs_excel_parser.py (tested two-phase loader).
Schema: migration_012_crs_module.sql + migration_014_crs_add_document_number.sql
must be applied before running.

Requirements:
    pip install pandas openpyxl python-calamine sqlalchemy pyyaml
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import sys
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import QueuePool

# Add repo root to path so etl.tasks.common is importable
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from etl.tasks.common import clean_string, load_config, get_db_engine_url, to_dt

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

# =============================================================================
# Constants (from scripts/crs_excel_parser.py — tested values)
# =============================================================================

MAIN_PATTERN = re.compile(
    r"^DOC_COMMENT_(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d+)_[A-Za-z]{3}\.xlsx$",
    re.IGNORECASE,
)
# Matches both versioned detail files (_A34_7) and Review_Comments (_A21_Review_Comments).
# re.IGNORECASE: tolerates review_comments / REVIEW_COMMENTS variants from vendors.
# \d+ instead of \d{2}: supports hypothetical revisions A100+ without pattern change.
# Version selection logic is in _select_detail_files(), not in this pattern.
DETAIL_PATTERN = re.compile(
    r"^(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d+)(?:_\d+|_Review_Comments)\.xlsx$",
    re.IGNORECASE,
)

COMMENT_COL_KEYWORDS: tuple[str, ...] = ("remark", "adura", "issue", "comment")

PROPERTY_COL_KEYWORDS: tuple[str, ...] = (
    "equipment property name",
    "equipment_property_name",
    "tag property name",
    "tag_property_name",
    "property name",
    "property_name",
)

# Keyword variants for the detail-sheet DOCUMENT_NUMBER column.
# This column references a project document containing the tag — distinct from
# the CRS file header number stored in crs_doc_number.
DOCUMENT_NUMBER_KEYWORDS: tuple[str, ...] = (
    "document number",
    "document_number",
    "doc number",
    "doc_number",
    "doc no",
    "doc_no",
)

FROM_TAG_KEYWORDS: tuple[str, ...] = (
    "from tag",
    "from_tag",
    "from tag name",
    "from_tag_name",
)

TO_TAG_KEYWORDS: tuple[str, ...] = (
    "to tag",
    "to_tag",
    "to tag name",
    "to_tag_name",
)

SKIP_SHEETS: set[str] = {"comment_sheet"}

HASH_EXCLUDE_FIELDS = {
    "sync_timestamp",
    "crs_file_timestamp",
    "llm_response_timestamp",
    "response_approval_date",
    "llm_response",
    # Excluded: change independently of comment content (filesystem paths, vendor responses)
    "response_vendor",
    "detail_file",
    "crs_file_path",
}

MAX_WORKERS = 14
BATCH_SIZE  = 500


# =============================================================================
# Parsing helpers (from scripts/crs_excel_parser.py)
# =============================================================================

def _scalar(val: Any) -> Any:
    """Return scalar value from a pandas Series or raw cell value.

    Args:
        val: A cell value or pandas Series from df.iterrows().

    Returns:
        Scalar value, or None if empty / NaN.
    """
    if isinstance(val, pd.Series):
        return val.iloc[0] if not val.empty else None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val

def _norm_sheet(name: str) -> str:
    """Normalise sheet name: strip, lowercase, spaces to underscores."""
    return name.strip().lower().replace(" ", "_")


_REV_RE = re.compile(r"_A(\d+)", re.IGNORECASE)


def _revision_label(key: str) -> str:
    """Extract revision label from a document key string.

    Args:
        key: Document key, e.g. 'JDAW-...-019_A28'.

    Returns:
        Revision label like 'A28', or 'A00' if not found.
    """
    m = _REV_RE.search(key)
    return f"A{m.group(1)}" if m else "A00"


def _revision_number(key: str) -> int:
    """Extract numeric revision from a document key string.

    Args:
        key: Document key, e.g. 'JDAW-...-019_A28'.

    Returns:
        Integer revision number, or 0 if not found.
    """
    m = _REV_RE.search(key)
    return int(m.group(1)) if m else 0

def _dir_revision(path: Path) -> str | None:
    """Extract revision label from parent directory name.

    Returns uppercase revision label like 'A19', or None if parent dir
    does not contain a revision suffix (validation skipped).
    """
    m = _REV_RE.search(path.parent.name)
    return f"A{m.group(1).upper()}" if m else None

_DETAIL_VERSION_RE = re.compile(r"_A\d+_(\d+)\.xlsx$", re.IGNORECASE)
_REVIEW_COMMENTS_RE = re.compile(r"_A\d+_Review_Comments\.xlsx$", re.IGNORECASE)


def _detail_version(filename: str) -> int:
    """Extract trailing integer version from a detail filename.

    Args:
        filename: Detail Excel filename, e.g. 'JDAW-...-003_A34_10.xlsx'.

    Returns:
        Integer version suffix (e.g. 10), or 0 if not a versioned file.

    Examples:
        >>> _detail_version("JDAW-...-003_A34_10.xlsx")
        10
        >>> _detail_version("JDAW-...-003_A21_Review_Comments.xlsx")
        0
    """
    m = _DETAIL_VERSION_RE.search(filename)
    return int(m.group(1)) if m else 0


def _is_review_comments(filename: str) -> bool:
    """Return True if filename matches the _Review_Comments convention.

    Args:
        filename: Detail Excel filename.

    Returns:
        True for '_A##_Review_Comments.xlsx' filenames, False otherwise.

    Examples:
        >>> _is_review_comments("JDAW-...-003_A21_Review_Comments.xlsx")
        True
        >>> _is_review_comments("JDAW-...-003_A34_10.xlsx")
        False
    """
    return bool(_REVIEW_COMMENTS_RE.search(filename))


def _select_detail_files(candidates: list[Path]) -> tuple[list[Path], str]:
    """Apply priority rules to select relevant detail files from candidates.

    All candidates are assumed to already be in the correct directory
    (Rule 0 applied by caller).

    Priority:
      1. _Review_Comments file → return it exclusively
      2. Versioned files (_A##_N) → return only max-version file(s)
      3. Fallback → return all candidates

    Args:
        candidates: List of eligible detail file paths (same dir as master).

    Returns:
        Tuple of (selected_paths, reason_string) for logging.
    """
    if not candidates:
        return [], "no candidates"

    # Priority 1: Review_Comments wins unconditionally
    review_files = [p for p in candidates if _is_review_comments(p.name)]
    if review_files:
        return review_files, "review_comments"

    # Priority 2: versioned files → keep only max-version
    versioned = [(p, _detail_version(p.name)) for p in candidates
                 if _detail_version(p.name) > 0]
    if versioned:
        max_ver = max(v for _, v in versioned)
        selected = [p for p, v in versioned if v == max_ver]
        return selected, f"max_version={max_ver}"

    # Priority 3: fallback (unusual naming — include all)
    return candidates, "fallback_all"


def _expand_merged_cells(ws) -> dict[tuple[int, int], object]:
    """Build a flat (row, col) -> value map with merged cells expanded.

    Openpyxl only exposes the top-left value for merged ranges. This function
    propagates that value to every cell in each merged region so callers can
    index any (row, col) directly.

    Args:
        ws: An openpyxl Worksheet object.

    Returns:
        Dict mapping (row, col) 1-based tuples to cell values.
    """
    cell_map: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for cell in row:
            cell_map[(cell.row, cell.column)] = cell.value
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = cell_map.get((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_map[(r, c)] = top_left_val
    return cell_map


def _build_two_row_header(
    cell_map: dict[tuple[int, int], object],
    row1: int,
    row2: int,
    max_col: int,
) -> list[str]:
    """Combine two header rows into a single deduplicated column name list.

    CRS main files use rows 6 and 7 as a split header. Values from both rows
    are joined with '_'. Duplicate names are disambiguated with a numeric suffix.

    Args:
        cell_map: Expanded cell map from _expand_merged_cells().
        row1: First header row index (1-based).
        row2: Second header row index (1-based).
        max_col: Number of columns to process.

    Returns:
        List of column name strings, upper-cased and sanitised.
    """
    headers = []
    for col in range(1, max_col + 1):
        v1 = str(cell_map.get((row1, col), "") or "").strip()
        v2 = str(cell_map.get((row2, col), "") or "").strip()
        if v1 and v2 and v1 != v2:
            name = f"{v1}_{v2}"
        elif v1:
            name = v1
        elif v2:
            name = v2
        else:
            name = f"COL_{col}"
        name = re.sub(r"[\s/\\().,-]+", "_", name).upper().strip("_")
        headers.append(name)
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def _find_comment_column(ws, data_start_row: int) -> int | None:
    """Detect the vertically-merged comment column in a detail sheet.

    The group comment in CRS detail files is stored as a tall merged cell
    spanning multiple data rows in a single column. This function finds the
    column with the largest vertical span starting at data_start_row.

    Args:
        ws: An openpyxl Worksheet object.
        data_start_row: Row index where data starts (usually 2).

    Returns:
        0-based column index of the merged comment column, or None if not found.
    """
    best_span: int = 1
    best_col:  int | None = None
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row_range = merged_range.bounds
        if min_col != max_col:
            continue
        if min_row != data_start_row:
            continue
        span = max_row_range - min_row + 1
        if span > best_span:
            best_span = span
            best_col  = min_col - 1
    return best_col


SheetData = tuple[pd.DataFrame, str | None, str | None]


def _load_detail_file_impl(path: Path) -> dict[str, SheetData]:
    """Load all sheets from a CRS detail Excel file (two-phase: openpyxl + calamine).

    Phase 1 (openpyxl): filter hidden sheets, detect merged comment column
    and extract its value per visible sheet.
    Phase 2 (calamine): read data rows efficiently; drop the merged column, filter
    empty rows, and detect fallback comment column if no merged cell was found.

    Args:
        path: Path to the detail Excel file.

    Returns:
        Dict mapping normalised sheet name to (DataFrame, comment_value, fallback_col).
        Empty dict if the file cannot be opened.
    """
    result: dict[str, SheetData] = {}
    DATA_START_ROW = 2

    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        log.error("Cannot open detail file %s: %s", path.name, exc)
        return result

    sheet_names = list(wb.sheetnames)

    # Filter hidden/veryHidden sheets — they contain service data, not CRS rows
    visible_sheet_names = [
        s for s in sheet_names
        if wb[s].sheet_state == "visible"
    ]
    if len(visible_sheet_names) < len(sheet_names):
        hidden = [s for s in sheet_names if s not in visible_sheet_names]
        log.debug(
            "Detail file %s: skipping %d hidden sheet(s): %s",
            path.name, len(hidden), ", ".join(hidden),
        )
    sheet_names = visible_sheet_names

    sheet_meta: dict[str, tuple[int | None, str | None]] = {}

    for sheet in sheet_names:
        sheet_key = _norm_sheet(sheet)
        if sheet_key in SKIP_SHEETS:
            continue
        if sheet_key in sheet_meta:
            log.warning(
                "Detail file %s: normalised sheet name collision '%s' "
                "(original: '%s') — skipping duplicate.",
                path.name, sheet_key, sheet,
            )
            continue
        ws = wb[sheet]
        comment_col_idx = _find_comment_column(ws, data_start_row=DATA_START_ROW)
        comment_value: str | None = None
        if comment_col_idx is not None:
            raw = ws.cell(row=DATA_START_ROW, column=comment_col_idx + 1).value
            comment_value = str(raw).strip() if raw is not None else None
            comment_value = comment_value or None
        sheet_meta[sheet_key] = (comment_col_idx, comment_value)

    wb.close()

    try:
        xl = pd.ExcelFile(path, engine="calamine")
    except Exception as exc:
        log.error("Cannot open detail file with calamine %s: %s", path.name, exc)
        return result

    for sheet in sheet_names:
        sheet_key = _norm_sheet(sheet)
        if sheet_key not in sheet_meta:
            continue
        comment_col_idx, comment_value = sheet_meta[sheet_key]
        try:
            df = pd.read_excel(xl, sheet_name=sheet, engine="calamine", dtype=str, na_filter=False)
        except Exception:
            continue
        df = df[~df.apply(lambda r: all(v == "" for v in r), axis=1)].copy()
        if df.empty:
            continue
        if comment_col_idx is not None and comment_col_idx < len(df.columns):
            df = df.drop(columns=df.columns[comment_col_idx])
        fallback_comment_col: str | None = None
        if comment_col_idx is None:
            for col in df.columns:
                if any(kw in col.lower() for kw in COMMENT_COL_KEYWORDS):
                    fallback_comment_col = col
                    break
        if fallback_comment_col is not None:
            df = df[df[fallback_comment_col].str.strip() != ""].copy()
            df = df.reset_index(drop=True)
            if df.empty:
                continue
        result[sheet_key] = (df, comment_value, fallback_comment_col)

    return result


_detail_file_cache: dict[Path, dict] = {}
_cache_lock = threading.Lock()


def _load_detail_file(path: Path) -> dict[str, SheetData]:
    # Fast path: dict lookup is atomic in CPython under GIL
    if path in _detail_file_cache:
        return _detail_file_cache[path]
    # Slow path: acquire lock to prevent duplicate parsing in thread pool
    with _cache_lock:
        if path not in _detail_file_cache:
            _detail_file_cache[path] = _load_detail_file_impl(path)
    return _detail_file_cache[path]

def _alphanum(s: str) -> str:
    """Strip all non-alphanumeric chars and lowercase. 'Master Doc' == 'MasterDoc'."""
    return re.sub(r"[^a-z0-9]", "", s.lower())

_EQUIP_PREFIX_RE = re.compile(r"^Equip_(.+)$", re.IGNORECASE)
def _extract_tag_from_equipment(val: Any) -> str | None:
    """
    Extract tag name from equipment number field.
    'Equip_ESB1_BUSCABLE1_0101' → 'ESB1_BUSCABLE1_0101'
    'Equip_JDA-01MV-0047'       → 'JDA-01MV-0047'
    Plain value without prefix  → returned as-is (fallback)
    """
    s = clean_string(val)
    if not s:
        return None
    m = _EQUIP_PREFIX_RE.match(s)
    return m.group(1) if m else s

_TAG_COL_EXACT_RE = re.compile(r"^tag[_\s]*(name|number|no\.?|num|id)?$", re.IGNORECASE)
_EQUIP_COL_RE     = re.compile(r"equipment[\s_]*(number|no\.?|num)", re.IGNORECASE)
_EQUIP_EXCLUDE_RE = re.compile(r"(serial|manufacturer|model|part)",   re.IGNORECASE)


def _find_tag_col(columns: list[str]) -> tuple[str | None, bool]:
    """Identify the tag or equipment number column in a detail sheet.

    Args:
        columns: List of column names from the detail sheet DataFrame.

    Returns:
        Tuple of (column_name, is_equipment_col). column_name is None if not found.
        is_equipment_col is True when the column holds equipment numbers (Equip_ prefix).
    """
    # Priority 1: explicit tag column (excludes property columns)
    for c in columns:
        if _TAG_COL_EXACT_RE.search(c) and "property" not in c.lower():
            return c, False

    # Priority 2: equipment number column as fallback
    for c in columns:
        if _EQUIP_COL_RE.search(c) and not _EQUIP_EXCLUDE_RE.search(c):
            return c, True

    return None, False


def find_matching_sheet(
    comment_text: str,
    sheets: dict[str, SheetData],
) -> tuple[str | None, pd.DataFrame | None, str | None, str | None]:
    """Match a group comment text to the best-fitting detail sheet.

    Uses normalised alphanumeric comparison. Longer sheet names are tried first
    to prefer more specific matches.

    Args:
        comment_text: GROUP_COMMENT value from the main CRS file.
        sheets: Dict mapping normalised sheet name to SheetData tuple.

    Returns:
        4-tuple (sheet_key, df, comment_value, fallback_col), or
        (None, None, None, None) if no match found.
    """
    comment_norm = _alphanum(comment_text)

    # Try longest sheet name first to prefer more specific matches
    for sheet_key, (df, comment_val, fallback_col) in sorted(
        sheets.items(), key=lambda x: len(x[0]), reverse=True
    ):
        if _alphanum(sheet_key) in comment_norm:
            return sheet_key, df, comment_val, fallback_col

    return None, None, None, None


def _report_orphans(orphan_sheets: list[dict]) -> None:
    """Log a summary of detail sheets that had no matching group comment.

    Args:
        orphan_sheets: List of orphan dicts from process_key().
    """
    if not orphan_sheets:
        return

    # Group by detail file for a compact report
    by_file: dict[str, list[dict]] = {}
    for o in orphan_sheets:
        by_file.setdefault(o["detail_file"], []).append(o)

    log.warning("=" * 60)
    log.warning("ORPHAN SHEETS REPORT — %d sheet(s) in %d file(s) not loaded to DB",
                len(orphan_sheets), len(by_file))
    log.warning("=" * 60)

    for detail_file, entries in sorted(by_file.items()):
        # Show all sheets per file — matched and unmatched
        e0 = entries[0]
        log.warning("")
        log.warning("  Detail file : %s", detail_file)
        log.warning("  Main file   : %s", e0["main_file"])
        log.warning("  All sheets  : %s", ", ".join(e0["available_sheets"]))
        log.warning("  Matched     : %s", ", ".join(e0["matched_sheets"]) or "— none —")
        log.warning("  ORPHAN(S)   :")
        for o in entries:
            log.warning("    ✗ sheet='%s'  rows=%d", o["sheet_key"], o["rows"])
            # Hint: compare this normalised form against GROUP_COMMENT values
            log.warning("      -> sheet normalized: '%s'", _alphanum(o["sheet_key"]))

    log.warning("")
    log.warning("ACTION REQUIRED: Check that GROUP_COMMENT text in main file")
    log.warning("  contains sheet name (after stripping spaces/special chars).")
    log.warning("=" * 60)

def _report_duplicates(dup_by_file: dict[str, int], total_raw: int, total_loaded: int) -> None:
    """Log a summary of duplicate records skipped during prepare_crs_records.

    Args:
        dup_by_file: Dict mapping source_file name to duplicate count.
        total_raw: Total raw records before deduplication.
        total_loaded: Records remaining after deduplication.
    """
    total_dups = sum(dup_by_file.values())
    log.info(
        "Prepared %d/%d DB records — %d duplicate(s) skipped.",
        total_loaded, total_raw, total_dups,
    )
    if dup_by_file:
        for src, cnt in sorted(dup_by_file.items(), key=lambda x: -x[1]):
            log.debug("  Duplicates in %s: %d", src, cnt)

def parse_main_file(path: Path) -> tuple[dict | None, pd.DataFrame | None]:
    """Parse a CRS main file and extract metadata + group comments DataFrame.

    Uses openpyxl for merged-cell-aware header reading, then calamine for fast
    data row loading. Skips files with RETURN_CODE == '1' (already responded).

    Args:
        path: Path to the DOC_COMMENT_* main Excel file.

    Returns:
        Tuple of (metadata_dict, df_comments), or (None, None) on error or skip.
        df_comments has columns GROUP_COMMENT and RESPONSE.
        metadata_dict contains DOC_NUMBER, REVISION, RETURN_CODE,
        TRANSMITTAL_NUMBER, TRANSMITTAL_DATE, SOURCE_FILE, CRS_FILE_PATH.
    """
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.active
    except Exception as exc:
        log.error("Cannot open %s: %s", path.name, exc)
        return None, None

    cell_map = _expand_merged_cells(ws)
    max_col  = ws.max_column or 0
    wb.close()

    def g(row: int, col: int) -> Any:
        val = cell_map.get((row, col))
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        return val

    metadata = {
        "DOC_NUMBER":         g(4, 3),
        "REVISION":           g(3, 6),
        "RETURN_CODE":        g(4, 6),
        "TRANSMITTAL_NUMBER": g(5, 3),
        "TRANSMITTAL_DATE":   g(5, 6),
        "SOURCE_FILE":        path.name,
        "CRS_FILE_PATH":      str(path),
    }

    rc = str(metadata["RETURN_CODE"]).strip().split(".")[0]
    if rc == "1":
        log.info("Skipping %s — Return Code 1.", path.name)
        return None, None

    if max_col < 3:
        return metadata, None

    headers = _build_two_row_header(cell_map, row1=6, row2=7, max_col=max_col)

    try:
        df_raw = pd.read_excel(path, header=None, skiprows=7, engine="calamine", dtype=str, na_filter=False)
    except Exception as exc:
        log.error("Cannot read data rows of %s: %s", path.name, exc)
        return metadata, None

    if df_raw.empty:
        return metadata, None

    actual_cols = df_raw.shape[1]
    if len(headers) < actual_cols:
        headers += [f"COL_{i}" for i in range(len(headers), actual_cols)]
    else:
        headers = headers[:actual_cols]
    df_raw.columns = headers

    col_c_name = headers[2] if len(headers) > 2 else None
    col_f_name = headers[5] if len(headers) > 5 else None
    if col_c_name is None:
        return metadata, None

    df_comments = pd.DataFrame()
    df_comments["GROUP_COMMENT"] = df_raw[col_c_name]
    df_comments["RESPONSE"]      = df_raw[col_f_name] if col_f_name else ""

    col_a_name   = headers[0] if headers else None
    mask_a_empty = (
        df_raw[col_a_name].str.strip() == ""
        if col_a_name else pd.Series([True] * len(df_raw))
    )
    mask_c_empty = df_comments["GROUP_COMMENT"].str.strip() == ""
    df_comments = df_comments[~(mask_a_empty & mask_c_empty)].copy()
    df_comments = df_comments[df_comments["GROUP_COMMENT"].str.strip() != ""].reset_index(drop=True)
    df_comments["RESPONSE"] = df_comments["RESPONSE"].replace("", None)

    return metadata, df_comments


def process_key(
    key: str,
    main_path: Path,
    related_paths: tuple[Path, ...],
) -> tuple[list[dict], list[dict], int]:
    """Parse one main CRS file and all its related detail files into raw records.

    For each group comment in the main file, finds the matching detail sheet,
    then emits one record per data row in that sheet. Tracks orphan sheets
    (detail sheets with no matching group comment) and duplicate comment IDs.

    Args:
        key: Document key identifying this CRS set (e.g. 'JDAW-...-019_A28').
        main_path: Path to the DOC_COMMENT_* main Excel file.
        related_paths: Tuple of paths to related detail Excel files.

    Returns:
        3-tuple of (records, orphan_sheets, dup_count).
    """
    records:      list[dict] = []
    matched_keys: set[str]   = set()
    seen_ids:     set[str]   = set()
    dup_count:    int        = 0

    metadata, df_comments = parse_main_file(main_path)
    if metadata is None or df_comments is None:
        return records, [], 0

    all_detail_sheets: dict[Path, dict[str, SheetData]] = {
        p: _load_detail_file(p) for p in related_paths
    }

    for _, row in df_comments.iterrows():
        comment_text  = str(row["GROUP_COMMENT"]).strip()
        response_text = _scalar(row.get("RESPONSE"))
        found_detail  = False

        for detail_path, sheets in all_detail_sheets.items():
            sheet_key, df_sheet, comment_val, fallback_col = find_matching_sheet(comment_text, sheets)
            if df_sheet is None:
                continue

            found_detail = True
            matched_keys.add(f"{detail_path.name}::{sheet_key}")

            # Resolve columns once per sheet before row loop
            tag_col, is_equip_col = _find_tag_col(list(df_sheet.columns))
            prop_col = next(
                (c for c in df_sheet.columns
                 if any(
                     kw.replace(" ", "_") in c.lower().replace(" ", "_")
                     for kw in PROPERTY_COL_KEYWORDS
                 )),
                None,
            )
            # DOCUMENT_NUMBER column is present only in some detail sheets
            doc_num_col = next(
                (c for c in df_sheet.columns
                 if any(
                     kw.replace(" ", "_") in c.lower().replace(" ", "_")
                     for kw in DOCUMENT_NUMBER_KEYWORDS
                 )),
                None,
            )
            # FROM_TAG / TO_TAG — present only in directional detail sheets
            from_tag_col = next(
                (c for c in df_sheet.columns
                 if any(
                     kw.replace(" ", "_") in c.lower().replace(" ", "_")
                     for kw in FROM_TAG_KEYWORDS
                 )),
                None,
            )
            to_tag_col = next(
                (c for c in df_sheet.columns
                 if any(
                     kw.replace(" ", "_") in c.lower().replace(" ", "_")
                     for kw in TO_TAG_KEYWORDS
                 )),
                None,
            )

            for _, d_row in df_sheet.iterrows():
                # Extract tag name — strip Equip_ prefix when column holds equipment numbers
                raw_tag  = _scalar(d_row[tag_col]) if tag_col else None
                tag_name = (
                    _extract_tag_from_equipment(raw_tag)
                    if is_equip_col
                    else clean_string(raw_tag)
                )

                prop_name = _scalar(d_row[prop_col]) if prop_col else None
                if not prop_name or str(prop_name).strip() == "":
                    prop_name = "Not Applicable"

                # document_number references the project document, not the CRS file
                doc_number_ref = _scalar(d_row[doc_num_col]) if doc_num_col else None
                if not doc_number_ref or str(doc_number_ref).strip() == "":
                    doc_number_ref = "Not Applicable"

                from_tag = clean_string(_scalar(d_row[from_tag_col])) if from_tag_col else None
                to_tag   = clean_string(_scalar(d_row[to_tag_col]))   if to_tag_col   else None

                row_comment = comment_val
                if row_comment is None and fallback_col and fallback_col in d_row.index:
                    row_comment = _scalar(d_row[fallback_col]) or None

                records.append({
                    "DOC_NUMBER":          metadata.get("DOC_NUMBER"),
                    "REVISION":            metadata.get("REVISION"),
                    "RETURN_CODE":         metadata.get("RETURN_CODE"),
                    "TRANSMITTAL_NUMBER":  metadata.get("TRANSMITTAL_NUMBER"),
                    "TRANSMITTAL_DATE":    metadata.get("TRANSMITTAL_DATE"),
                    "TAG_NAME":            tag_name,
                    "PROPERTY_NAME":       prop_name,
                    "DOCUMENT_NUMBER_REF": doc_number_ref,
                    "FROM_TAG":            from_tag,
                    "TO_TAG":              to_tag,
                    "GROUP_COMMENT":       comment_text,
                    "RESPONSE":            response_text,
                    "SOURCE_FILE":         metadata.get("SOURCE_FILE"),
                    "DETAIL_FILE":         detail_path.name,
                    "DETAIL_SHEET":        sheet_key,
                    "COMMENT":             row_comment,
                    "CRS_FILE_PATH":       metadata.get("CRS_FILE_PATH"),
                })

                prop_key_chk = (
                    prop_name
                    if prop_name and prop_name.upper() != "NOT APPLICABLE"
                    else ""
                )
                doc_num_chk = (
                    doc_number_ref
                    if doc_number_ref and doc_number_ref.upper() != "NOT APPLICABLE"
                    else ""
                )
                _cid = (
                    f"{metadata.get('DOC_NUMBER', '')}|{comment_text}|{sheet_key}"
                    f"|{tag_name or ''}|{row_comment or ''}|{prop_key_chk}|{doc_num_chk}"
                    f"|{from_tag or ''}|{to_tag or ''}"
                )
                if _cid in seen_ids:
                    dup_count += 1
                else:
                    seen_ids.add(_cid)

            break  # stop after first matching sheet per group comment

        if not found_detail:
            records.append({
                "DOC_NUMBER":          metadata.get("DOC_NUMBER"),
                "REVISION":            metadata.get("REVISION"),
                "RETURN_CODE":         metadata.get("RETURN_CODE"),
                "TRANSMITTAL_NUMBER":  metadata.get("TRANSMITTAL_NUMBER"),
                "TRANSMITTAL_DATE":    metadata.get("TRANSMITTAL_DATE"),
                "TAG_NAME":            None,
                "PROPERTY_NAME":       "Not Applicable",
                "DOCUMENT_NUMBER_REF": "Not Applicable",
                "FROM_TAG":            None,
                "TO_TAG":              None,
                "GROUP_COMMENT":       comment_text,
                "RESPONSE":            response_text,
                "SOURCE_FILE":         metadata.get("SOURCE_FILE"),
                "DETAIL_FILE":         None,
                "DETAIL_SHEET":        None,
                "COMMENT":             "No matching detail sheet found",
                "CRS_FILE_PATH":       metadata.get("CRS_FILE_PATH"),
            })

    orphan_sheets: list[dict] = []
    for detail_path, sheets in all_detail_sheets.items():
        available_sheets   = list(sheets.keys())
        matched_for_file   = [
            mk.split("::")[1]
            for mk in matched_keys
            if mk.startswith(detail_path.name + "::")
        ]
        unmatched_for_file = [sk for sk in available_sheets if sk not in matched_for_file]

        for sk in unmatched_for_file:
            df_orphan, _, _ = sheets[sk]
            orphan_sheets.append({
                "detail_file":      detail_path.name,
                "sheet_key":        sk,
                "rows":             len(df_orphan),
                "main_file":        main_path.name,
                "available_sheets": available_sheets,
                "matched_sheets":   matched_for_file,
            })

    return records, orphan_sheets, dup_count


# =============================================================================
# DB operations
# =============================================================================

def discover_crs_files(
    root: Path,
) -> tuple[dict[str, Path], dict[str, list[Path]], list[str]]:
    """Scan CRS directory and classify files into main and detail sets.

    Detail file selection rules (per document key):
      0. Only files in the SAME directory as master are eligible (hard rule)
      1. _Review_Comments file wins if present
      2. Otherwise: max-version integer-suffixed file wins
      3. Fallback: keep all eligible

    Args:
        root: Root directory to scan recursively for .xlsx files.

    Returns:
        3-tuple of:
          - main_files: {document_key: path}
          - detail_files: {document_key: [selected_paths]}
          - rev_order: revision labels sorted A01 -> A99
    """
    main_files:  dict[str, Path]       = {}
    _all_detail: dict[str, list[Path]] = {}

    for path in root.rglob("*.xlsx"):
        if "_templates" in path.parts:
            continue
        if path.name.startswith("~$"):
            continue
        if "Surplus" in path.parts:
            log.debug("SKIP (Surplus): %s", path)
            continue
        log.debug("SCAN: %s | parts=%s", path.name, path.parts)
        name = path.name
        m = MAIN_PATTERN.match(name)
        if m:
            key = m.group(1).upper()
            file_rev = _revision_label(key)
            dir_rev  = _dir_revision(path)
            if dir_rev is not None and file_rev != dir_rev:
                log.warning(
                    "SKIP (rev mismatch) main file: %s\n"
                    "  file revision : %s\n"
                    "  dir  revision : %s\n"
                    "  path          : %s",
                    name, file_rev, dir_rev, path,
                )
                continue
            if key not in main_files:
                main_files[key] = path
            else:
                log.warning("Duplicate main key %s — keeping first.", key)
            continue
        d = DETAIL_PATTERN.match(name)
        if d:
            key = d.group(1).upper()
            file_rev = _revision_label(key)
            dir_rev  = _dir_revision(path)
            if dir_rev is not None and file_rev != dir_rev:
                log.warning(
                    "SKIP (rev mismatch) detail file: %s\n"
                    "  file revision : %s\n"
                    "  dir  revision : %s\n"
                    "  path          : %s",
                    name, file_rev, dir_rev, path,
                )
                continue
            _all_detail.setdefault(key, []).append(path)
            continue
        log.debug("Unmatched xlsx (not CRS): %s", path.relative_to(root))

    detail_files: dict[str, list[Path]] = {}

    for key, candidates in _all_detail.items():
        master_path = main_files.get(key)

        if master_path is None:
            log.warning(
                "Key %s: %d detail file(s) found but no master file — skipped.",
                key, len(candidates),
            )
            continue

        master_dir = master_path.parent.resolve()

        # Rule 0: same-directory filter (hard — subdirectory files are legacy)
        same_dir = [p for p in candidates if p.parent.resolve() == master_dir]
        sub_dir  = [p for p in candidates if p.parent.resolve() != master_dir]

        if sub_dir:
            for p in sub_dir:
                log.info(
                    "Key %s: excluded detail (dir mismatch):\n"
                    "  detail dir : %s\n"
                    "  master dir : %s",
                    key, p.parent.resolve(), master_dir,
                )

        if not same_dir:
            log.warning(
                "Key %s: all %d detail file(s) are in subdirectories — none eligible.",
                key, len(candidates),
            )
            continue

        # Rules 1–3: priority selection within same directory
        selected, reason = _select_detail_files(same_dir)
        excluded_local = [p for p in same_dir if p not in selected]

        if excluded_local:
            log.info(
                "Key %s: selected %d file(s) [%s], excluded %d older: %s",
                key, len(selected), reason,
                len(excluded_local),
                ", ".join(p.name for p in excluded_local),
            )
        else:
            log.debug("Key %s: selected %d file(s) [%s]: %s",
                      key, len(selected), reason,
                      ", ".join(p.name for p in selected))

        detail_files[key] = selected

    rev_order = sorted(
        {_revision_label(k) for k in main_files},
        key=lambda r: int(r[1:]) if r[1:].isdigit() else 0,
    )

    log.info(
        "Found %d main file(s) across %d revision(s): %s",
        len(main_files), len(rev_order), ", ".join(rev_order),
    )
    return main_files, detail_files, rev_order


def parse_all_files(
    main_files:   dict[str, Path],
    detail_files: dict[str, list[Path]],
    rev_order:    list[str],
    debug_rev:    str | None = None,
) -> tuple[list[dict], list[str]]:
    """Parse all CRS file sets in revision order using a thread pool.

    Files are processed A01 -> A99. Within each revision group, all document
    keys are processed in parallel via ThreadPoolExecutor.

    Args:
        main_files: {key: path} mapping from discover_crs_files().
        detail_files: {key: [paths]} mapping from discover_crs_files().
        rev_order: Sorted revision label list from discover_crs_files().
        debug_rev: If set, process only this revision (e.g. 'A28'). None = all.

    Returns:
        Tuple of (all_records, orphan_sheets).
    """
    # Group keys by revision for ordered processing
    groups: dict[str, list[str]] = {}
    for key in main_files:
        rev = _revision_label(key)
        groups.setdefault(rev, []).append(key)

    revs_to_process = [debug_rev] if debug_rev else rev_order

    all_records:   list[dict] = []
    orphan_sheets: list[dict] = []
    total_dups:    int        = 0

    for rev in revs_to_process:
        keys = sorted(groups.get(rev, []))  # sort within revision by document key
        if not keys:
            log.warning("Revision %s requested but no files found.", rev)
            continue

        log.info("── Revision %s: %d file(s) ──", rev, len(keys))

        work_items = [
            (key, main_files[key], tuple(detail_files.get(key, [])))
            for key in keys
        ]

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(process_key, key, main_path, related): key
                for key, main_path, related in work_items
            }
            for future in as_completed(futures):
                key = futures[future]
                try:
                    records, file_orphans, dup_count = future.result()
                    all_records.extend(records)
                    orphan_sheets.extend(file_orphans)
                    total_dups += dup_count
                    log.info("  ✓ %s — %d record(s), %d orphan(s), %d duplicate(s)", key, len(records), len(file_orphans), dup_count)
                except Exception as exc:
                    log.error("  ✗ %s failed: %s", key, exc)

    _detail_file_cache.clear()
    log.info(
        "Parsed %d total records, %d duplicate(s), %d orphan sheet(s).",
        len(all_records), total_dups, len(orphan_sheets),
    )
    return all_records, orphan_sheets


def prepare_crs_records(raw_records: list[dict], engine) -> list[dict]:
    """Enrich raw parsed records with FK resolution, row_hash and comment_id.

    Performs a single batched FK lookup for all unique tag names, then builds
    DB-ready dicts with deterministic comment_id (UUID5) and MD5 row_hash.
    Deduplicates records sharing the same comment_id within the batch.

    Args:
        raw_records: List of raw record dicts from parse_all_files().
        engine: SQLAlchemy engine for FK lookup queries.

    Returns:
        List of DB-ready dicts suitable for upsert_crs_records().
    """
    if not raw_records:
        return []

    # 1. Collect unique tag names needed for FK lookup
    needed_tags = {
        clean_string(r.get("TAG_NAME"))
        for r in raw_records
        if clean_string(r.get("TAG_NAME"))
    }

    # 2. Single batched FK lookup using ANY() — avoids N+1 queries
    with engine.connect() as conn:
        tag_lookup: dict[str, str] = {}
        if needed_tags:
            tag_lookup = {
                row.tag_name: str(row.id)
                for row in conn.execute(
                    text("""
                        SELECT id, tag_name
                        FROM project_core.tag
                        WHERE tag_name = ANY(:names)
                          AND object_status = 'Active'
                    """),
                    {"names": list(needed_tags)},
                )
            }

    log.info("FK lookup: %d/%d tags resolved.", len(tag_lookup), len(needed_tags))

    # 3. Build DB records
    db_records: list[dict] = []
    tag_miss = 0
    seen_comment_ids: set[str] = set()
    dup_by_file: dict[str, int] = {}

    for rec in raw_records:
        # Call clean_string once per field and reuse — avoids repeated calls
        crs_doc_number      = clean_string(rec.get("DOC_NUMBER")) or "UNKNOWN"
        tag_name            = clean_string(rec.get("TAG_NAME"))
        revision            = clean_string(rec.get("REVISION"))
        return_code         = clean_string(rec.get("RETURN_CODE"))
        transmittal_num     = clean_string(rec.get("TRANSMITTAL_NUMBER"))
        transmittal_date    = to_dt(rec.get("TRANSMITTAL_DATE"))
        group_comment       = clean_string(rec.get("GROUP_COMMENT")) or ""
        comment             = clean_string(rec.get("COMMENT")) or ""
        property_name       = clean_string(rec.get("PROPERTY_NAME"))
        document_number_ref = clean_string(rec.get("DOCUMENT_NUMBER_REF")) or "Not Applicable"
        from_tag            = clean_string(rec.get("FROM_TAG"))
        to_tag              = clean_string(rec.get("TO_TAG"))
        response_vendor     = clean_string(rec.get("RESPONSE"))
        source_file         = clean_string(rec.get("SOURCE_FILE")) or ""
        detail_file         = clean_string(rec.get("DETAIL_FILE"))
        detail_sheet        = clean_string(rec.get("DETAIL_SHEET"))
        crs_file_path       = clean_string(rec.get("CRS_FILE_PATH")) or ""

        # FK resolve — None when tag not found (logged as warning at end)
        tag_id = tag_lookup.get(tag_name) if tag_name else None
        if tag_name and not tag_id:
            tag_miss += 1

        # row_hash: MD5 over stable content fields (excludes timestamps and
        # filesystem-dependent paths that change independently of comment content)
        hash_source = {
            "crs_doc_number":      crs_doc_number,
            "tag_name":            tag_name or "",
            "revision":            revision or "",
            "return_code":         return_code or "",
            "transmittal_num":     transmittal_num or "",
            "transmittal_date":    str(transmittal_date) if transmittal_date else "",
            "group_comment":       group_comment,
            "comment":             comment,
            "property_name":       property_name or "",
            "document_number_ref": document_number_ref,
            "from_tag":            from_tag or "",
            "to_tag":              to_tag or "",
            "detail_sheet":        detail_sheet or "",
        }
        row_hash = hashlib.md5(
            json.dumps(hash_source, sort_keys=True).encode()
        ).hexdigest()

        # comment_id: deterministic UUID5 — stable across re-runs
        prop_key = (
            property_name
            if property_name and property_name.upper() != "NOT APPLICABLE"
            else ""
        )
        doc_num_key = (
            document_number_ref
            if document_number_ref and document_number_ref.upper() != "NOT APPLICABLE"
            else ""
        )
        comment_id = str(uuid.uuid5(
            uuid.NAMESPACE_URL,
            f"{crs_doc_number}|{group_comment}|{detail_sheet or ''}"
            f"|{tag_name or ''}|{comment}|{prop_key}|{doc_num_key}"
            f"|{from_tag or ''}|{to_tag or ''}",
        ))

        # Deduplicate within batch — guards against repeated rows in Excel source
        if comment_id in seen_comment_ids:
            src = clean_string(rec.get("SOURCE_FILE")) or "UNKNOWN"
            dup_by_file[src] = dup_by_file.get(src, 0) + 1
            continue
        seen_comment_ids.add(comment_id)

        db_records.append({
            "comment_id":         comment_id,
            "crs_doc_number":     crs_doc_number,
            "revision":           revision,
            "return_code":        return_code,
            "transmittal_number": transmittal_num,
            "transmittal_date":   transmittal_date,
            "group_comment":      group_comment,
            "comment":            comment,
            "tag_name":           tag_name,
            "tag_id":             tag_id,
            "property_name":      property_name,
            "document_number":    document_number_ref,
            "from_tag":           from_tag,
            "to_tag":             to_tag,
            "response_vendor":    response_vendor,
            "source_file":        source_file,
            "detail_file":        detail_file,
            "detail_sheet":       detail_sheet,
            "crs_file_path":      crs_file_path,
            "crs_file_timestamp": None,
            "status":             "RECEIVED",
            "object_status":      "Active",
            "row_hash":           row_hash,
        })

    if tag_miss:
        log.warning("FK miss — tag_name unresolved: %d (tag_id=NULL)", tag_miss)

    _report_duplicates(
        dup_by_file,
        total_raw=len(raw_records),
        total_loaded=len(db_records),
    )

    return db_records


def upsert_crs_records(engine, records: list[dict], run_id: str) -> dict[str, int]:
    """Upsert CRS comment records into audit_core.crs_comment in batches.

    Uses ON CONFLICT (comment_id) DO UPDATE with a row_hash guard to skip
    unchanged rows. Writes an audit entry to crs_comment_audit for every
    INSERT or UPDATE.

    Args:
        engine: SQLAlchemy engine.
        records: DB-ready dicts from prepare_crs_records().
        run_id: UUID string for the current run (written to audit table).

    Returns:
        Dict with keys 'inserted', 'updated', 'errors'.
    """
    stats = {"inserted": 0, "updated": 0, "errors": 0}
    if not records:
        return stats

    upsert_sql = text("""
        INSERT INTO audit_core.crs_comment (
            comment_id, crs_doc_number, revision, return_code,
            transmittal_number, transmittal_date,
            group_comment, comment, tag_name, tag_id, property_name,
            document_number, from_tag, to_tag,
            response_vendor, source_file, detail_file, detail_sheet,
            crs_file_path, crs_file_timestamp,
            status, object_status, row_hash, sync_timestamp
        ) VALUES (
            :comment_id, :crs_doc_number, :revision, :return_code,
            :transmittal_number, :transmittal_date,
            :group_comment, :comment, :tag_name, :tag_id, :property_name,
            :document_number, :from_tag, :to_tag,
            :response_vendor, :source_file, :detail_file, :detail_sheet,
            :crs_file_path, :crs_file_timestamp,
            :status, :object_status, :row_hash, now()
        )
        ON CONFLICT (comment_id) DO UPDATE SET
            group_comment      = EXCLUDED.group_comment,
            comment            = EXCLUDED.comment,
            tag_name           = EXCLUDED.tag_name,
            tag_id             = EXCLUDED.tag_id,
            property_name      = EXCLUDED.property_name,
            document_number    = EXCLUDED.document_number,
            from_tag           = EXCLUDED.from_tag,
            to_tag             = EXCLUDED.to_tag,
            response_vendor    = EXCLUDED.response_vendor,
            crs_file_timestamp = EXCLUDED.crs_file_timestamp,
            row_hash           = EXCLUDED.row_hash,
            sync_timestamp     = now()
        WHERE audit_core.crs_comment.row_hash != EXCLUDED.row_hash
        RETURNING id, xmax
    """)

    audit_sql = text("""
        INSERT INTO audit_core.crs_comment_audit (comment_id, change_type, snapshot, run_id)
        VALUES (:cid, :ct, CAST(:snap AS jsonb), :rid)
    """)

    for batch_start in range(0, len(records), BATCH_SIZE):
        batch = records[batch_start : batch_start + BATCH_SIZE]
        try:
            with engine.begin() as conn:
                audit_rows = []
                for rec in batch:
                    result = conn.execute(upsert_sql, rec)
                    row = result.fetchone()
                    if row is None:
                        continue
                    change_type = "INSERT" if row.xmax == 0 else "UPDATE"
                    if row.xmax == 0:
                        stats["inserted"] += 1
                    else:
                        stats["updated"] += 1
                    snap = {k: str(v) for k, v in rec.items() if v is not None}
                    audit_rows.append({
                        "cid": str(row.id),
                        "ct": change_type,
                        "snap": json.dumps(snap),
                        "rid": run_id,
                    })

                # One executemany instead of N execute:
                if audit_rows:
                    conn.execute(audit_sql, audit_rows)
        except Exception as exc:
            log.error("Batch %d–%d error: %s", batch_start, batch_start + len(batch), exc)
            stats["errors"] += len(batch)

    log.info("Upsert done: inserted=%d updated=%d errors=%d",
             stats["inserted"], stats["updated"], stats["errors"])
    return stats


# =============================================================================
# Main
# =============================================================================

def run(debug_mode: bool = False, debug_rev: str | None = None) -> None:
    run_id = str(uuid.uuid4())
    config  = load_config()
    db_url  = get_db_engine_url(config)

    crs_data_dir = config.get("storage", {}).get("crs_data_dir")
    if not crs_data_dir:
        log.error("storage.crs_data_dir missing from config/config.yaml")
        sys.exit(1)

    root = Path(crs_data_dir)
    if not root.exists():
        log.error("CRS data directory not found: %s", root)
        sys.exit(1)

    log.info("=== CRS Import | run_id=%s | debug=%s ===", run_id, debug_mode)

    engine = create_engine(
        db_url,
        poolclass=QueuePool,
        pool_size=14,        # = MAX_WORKERS
        max_overflow=4,
        pool_recycle=1800,   
        pool_pre_ping=True,  
    )

    # Fail-fast DB check
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
    except Exception as exc:
        log.error("DB connection failed: %s", exc)
        sys.exit(1)

    # Audit: INSERT on start
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO audit_core.sync_run_stats (run_id, target_table, start_time, source_file)
            VALUES (:rid, :tbl, :st, :sf)
        """), {"rid": run_id, "tbl": "audit_core.crs_comment", "st": datetime.now(), "sf": str(root)})

    # Step 1: Discover
    main_files, detail_files, rev_order = discover_crs_files(root)

    active_rev = None
    if debug_mode:
        active_rev = debug_rev or (rev_order[-1] if rev_order else None)
        log.info("DEBUG mode: processing revision %s only.", active_rev)

    all_records, orphan_sheets = parse_all_files(
        main_files, detail_files, rev_order,
        debug_rev=active_rev,
    )

    stats: dict[str, int] = {"inserted": 0, "updated": 0, "errors": 0}

    if all_records:
        # Step 3: Prepare (hash + FK resolution)
        db_records = prepare_crs_records(all_records, engine)

        # Step 4: Upsert
        stats = upsert_crs_records(engine, db_records, run_id)
    else:
        log.warning("No records parsed — nothing to upsert.")

    # Audit: UPDATE on completion
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE audit_core.sync_run_stats SET
                end_time        = :et,
                count_created   = :cr,
                count_updated   = :up,
                count_unchanged = :uc,
                count_deleted   = :dl,
                count_errors    = :er
            WHERE run_id = :rid
        """), {
            "et": datetime.now(),
            "cr": stats["inserted"], "up": stats["updated"],
            "uc": 0, "dl": 0, "er": stats["errors"],
            "rid": run_id,
        })

    engine.dispose()

    _report_orphans(orphan_sheets)

    log.info(
        "=== DONE | parsed=%d loaded=%d errors=%d orphans=%d ===",
        len(all_records), stats["inserted"] + stats["updated"],
        stats["errors"], len(orphan_sheets),
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import CRS comments into PostgreSQL")
    parser.add_argument("--debug",     action="store_true")
    parser.add_argument("--debug-rev", default=None,
                        help="Process only this revision group, e.g. A28")
    args = parser.parse_args()
    run(debug_mode=args.debug, debug_rev=args.debug_rev)
