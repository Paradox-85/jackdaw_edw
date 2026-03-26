"""
CRS Excel Parser with Database Integration
Reads CRS Excel files (DOC_COMMENT_* and JDAW_*.xlsx) and loads directly into PostgreSQL.

Extends original crs_excel_parser.py with:
  - Database write capability
  - Row hash calculation for SCD2 tracking
  - Comment ID generation
  - Batch upsert logic

Requirements:
    pip install pandas openpyxl python-calamine sqlalchemy psycopg2-binary
"""

import re
import logging
import threading
import hashlib
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text, Connection, Engine
from sqlalchemy.dialects.postgresql import insert

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers (from original crs_excel_parser.py)
# ---------------------------------------------------------------------------

def _scalar(val):
    """Extract a scalar from a value that may be a pd.Series."""
    if isinstance(val, pd.Series):
        return val.iloc[0] if not val.empty else None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def _norm_sheet(name: str) -> str:
    """Normalise sheet name: strip, lowercase, spaces → underscores."""
    return name.strip().lower().replace(" ", "_")


def _calculate_row_hash(row_dict: dict) -> str:
    """
    Calculate MD5 hash of row data (excluding timestamps and hashes).
    Used for SCD2 change detection.
    """
    # Exclude these fields from hashing
    exclude_fields = {
        'sync_timestamp', 'row_hash', 'llm_response_timestamp', 
        'validation_timestamp', 'response_approval_date', 
        'crs_file_timestamp', 'llm_response'
    }
    
    content = "".join(
        str(v) for k, v in sorted(row_dict.items()) 
        if k not in exclude_fields and v is not None
    )
    return hashlib.md5(content.encode()).hexdigest()


def _parse_date(val: str) -> Optional[str]:
    """Parse date string to YYYY-MM-DD format. Returns None if unparseable."""
    if not val or pd.isna(val):
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    
    # Try common formats
    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
        try:
            dt = datetime.strptime(val_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    log.warning(f"Could not parse date: {val_str}")
    return None


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SOURCE_DIR = Path("/mnt/shared-data/ram-user/Jackdaw/EIS-Data")

MAX_WORKERS = 6
BATCH_SIZE = 500  # Insert batch size

SKIP_SHEETS = {_norm_sheet("Comment Sheet")}

MAIN_PATTERN = re.compile(
    r"^DOC_COMMENT_(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d{2})_[A-Z]{3}\.xlsx$"
)
DETAIL_PATTERN = re.compile(
    r"^(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d{2})(?:_\d+|_Review_Comments)\.xlsx$"
)

COMMENT_COL_KEYWORDS = ("remark", "adura", "issue", "comment")
PROPERTY_COL_KEYWORDS = ("equipment property name", "tag property name", "property name")

OUTPUT_COLS = [
    "DOC_NUMBER", "REVISION", "RETURN_CODE", "TRANSMITTAL_NUMBER",
    "TRANSMITTAL_DATE", "TAG_NAME", "PROPERTY_NAME", "GROUP_COMMENT",
    "RESPONSE", "SOURCE_FILE", "DETAIL_FILE", "DETAIL_SHEET", "COMMENT",
]


# ---------------------------------------------------------------------------
# File discovery (from original)
# ---------------------------------------------------------------------------

def discover_files(root: Path) -> tuple[dict[str, Path], dict[str, list[Path]]]:
    """Recursively scan root, skip _templates at any depth."""
    main_files: dict[str, Path] = {}
    detail_files: dict[str, list[Path]] = {}

    for path in root.rglob("*.xlsx"):
        if "_templates" in path.parts:
            continue

        name = path.name
        m = MAIN_PATTERN.match(name)
        if m:
            key = m.group(1)
            if key in main_files:
                log.warning("Duplicate main file key %s — keeping: %s", key, main_files[key])
            else:
                main_files[key] = path
            continue

        d = DETAIL_PATTERN.match(name)
        if d:
            detail_files.setdefault(d.group(1), []).append(path)

    log.info("Found %d main file(s), %d detail key(s).", len(main_files), len(detail_files))
    return main_files, detail_files


# ---------------------------------------------------------------------------
# Merged-cell resolution (from original)
# ---------------------------------------------------------------------------

def _expand_merged_cells(ws) -> dict[tuple[int, int], object]:
    cell_map: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for cell in row:
            cell_map[(cell.row, cell.column)] = cell.value

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = cell_map.get((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_map[(r, c)] = top_left_val

    return cell_map


def _build_two_row_header(
    cell_map: dict, row1: int, row2: int, max_col: int
) -> list[str]:
    """Combine two header rows into UPPER_SNAKE column names. Deduplicates."""
    headers = []
    for col in range(1, max_col + 1):
        v1 = str(cell_map.get((row1, col), "") or "").strip()
        v2 = str(cell_map.get((row2, col), "") or "").strip()

        if v1 and v2 and v1 != v2:
            name = f"{v1}_{v2}"
        elif v1:
            name = v1
        elif v2:
            name = v2
        else:
            name = f"COL_{col}"

        name = re.sub(r"[\s/\\().,-]+", "_", name).upper().strip("_")
        headers.append(name)

    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def _find_comment_column(ws, data_start_row: int) -> Optional[int]:
    """Detect a vertically-merged single-column comment cell."""
    best_span: int = 1
    best_col: Optional[int] = None

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row_range = merged_range.bounds

        if min_col != max_col:
            continue
        if min_row != data_start_row:
            continue

        span = max_row_range - min_row + 1
        if span > best_span:
            best_span = span
            best_col = min_col - 1

    return best_col


# ---------------------------------------------------------------------------
# Main-file parser (from original, simplified)
# ---------------------------------------------------------------------------

def parse_main_file(path: Path) -> tuple[Optional[dict], Optional[pd.DataFrame]]:
    """Extract header metadata and comment table from a DOC_COMMENT_* file."""
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.active
    except Exception as exc:
        log.error("Cannot open workbook %s: %s", path.name, exc)
        return None, None

    cell_map = _expand_merged_cells(ws)
    max_col = ws.max_column or 0
    wb.close()

    def g(row: int, col: int):
        val = cell_map.get((row, col))
        if val is None:
            return None
        return str(val).strip() if val else None

    # Extract header (rows 1-5)
    metadata = {
        "DOC_NUMBER": g(1, 1),
        "REVISION": g(2, 1),
        "RETURN_CODE": g(3, 1),
        "TRANSMITTAL_NUMBER": g(4, 1),
        "TRANSMITTAL_DATE": g(5, 1),
        "SOURCE_FILE": path.name,
    }

    # Read data from row 8 onwards
    try:
        df = pd.read_excel(path, header=None, dtype=str, na_filter=False)
        if len(df) < 8:
            return metadata, None
        
        df = df.iloc[7:].reset_index(drop=True)
        df.columns = _build_two_row_header(cell_map, 7, 8, max_col) if max_col > 0 else range(len(df.columns))
        
    except Exception as exc:
        log.error("Cannot read data from %s: %s", path.name, exc)
        return metadata, None

    return metadata, df


# ---------------------------------------------------------------------------
# Detail-file parser (simplified from original)
# ---------------------------------------------------------------------------

_detail_file_cache: dict[Path, dict] = {}
_cache_lock = threading.Lock()


def _load_detail_file_impl(path: Path) -> dict:
    """Load detail sheets from JDAW_* file."""
    result = {}
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        log.error("Cannot open detail file %s: %s", path.name, exc)
        return result

    for sheet in wb.sheetnames:
        sheet_key = _norm_sheet(sheet)
        if sheet_key in SKIP_SHEETS:
            continue

        try:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str, na_filter=False)
            result[sheet_key] = df
        except Exception as exc:
            log.error("Cannot read sheet %s from %s: %s", sheet, path.name, exc)
            continue

    wb.close()
    return result


def _load_detail_file(path: Path) -> dict:
    """Thread-safe cache wrapper."""
    if path in _detail_file_cache:
        return _detail_file_cache[path]
    with _cache_lock:
        if path not in _detail_file_cache:
            _detail_file_cache[path] = _load_detail_file_impl(path)
    return _detail_file_cache[path]


def find_matching_sheet(comment_text: str, sheets: dict) -> tuple[Optional[str], Optional[pd.DataFrame]]:
    """Find matching detail sheet for a comment."""
    text_norm = comment_text.lower().replace(" ", "_")
    for sheet_key, df in sheets.items():
        if sheet_key in text_norm:
            return sheet_key, df
    return None, None


# ---------------------------------------------------------------------------
# Per-key processing unit (from original, adapted for DB)
# ---------------------------------------------------------------------------

def process_key(
    key: str,
    main_path: Path,
    related_paths: tuple[Path, ...],
) -> list[dict]:
    """Returns list of records with mapped column names for DB insert."""
    records: list[dict] = []

    metadata, df_comments = parse_main_file(main_path)
    if metadata is None or df_comments is None:
        return records

    all_detail_sheets: dict[Path, dict] = {
        p: _load_detail_file(p) for p in related_paths
    }

    for idx, row in df_comments.iterrows():
        comment_text = str(row.get("GROUP_COMMENT", "")).strip()
        response_text = _scalar(row.get("RESPONSE"))
        found_detail = False

        for detail_path, sheets in all_detail_sheets.items():
            sheet_name, df_sheet = find_matching_sheet(comment_text, sheets)
            if df_sheet is None:
                continue

            found_detail = True

            # Find TAG_NAME column
            tag_col = next(
                (c for c in df_sheet.columns
                 if "tag" in c.lower() and "property" not in c.lower()), None
            )

            # Find PROPERTY_NAME column
            prop_col = next(
                (c for c in df_sheet.columns
                 if any(kw in c.lower() for kw in PROPERTY_COL_KEYWORDS)), None
            )

            for _, d_row in df_sheet.iterrows():
                tag_name = _scalar(d_row[tag_col]) if tag_col else None
                prop_name = _scalar(d_row[prop_col]) if prop_col else None
                if not prop_name or str(prop_name).strip() == "":
                    prop_name = "Not Applicable"

                rec = {
                    "doc_number": metadata.get("DOC_NUMBER"),
                    "revision": metadata.get("REVISION"),
                    "return_code": metadata.get("RETURN_CODE"),
                    "transmittal_number": metadata.get("TRANSMITTAL_NUMBER"),
                    "transmittal_date": _parse_date(metadata.get("TRANSMITTAL_DATE")),
                    "tag_name": tag_name,
                    "property_name": prop_name,
                    "group_comment": comment_text,
                    "response_vendor": response_text,
                    "source_file": metadata.get("SOURCE_FILE"),
                    "detail_file": detail_path.name,
                    "detail_sheet": sheet_name,
                    "comment": str(d_row.get(_scalar([c for c in d_row.index if "comment" in c.lower()][0]) if any("comment" in c.lower() for c in d_row.index) else ""))).strip() or "No specific detail",
                    "crs_file_path": str(main_path),
                    "crs_file_timestamp": datetime.fromtimestamp(main_path.stat().st_mtime),
                }
                records.append(rec)

            break

        if not found_detail:
            rec = {
                "doc_number": metadata.get("DOC_NUMBER"),
                "revision": metadata.get("REVISION"),
                "return_code": metadata.get("RETURN_CODE"),
                "transmittal_number": metadata.get("TRANSMITTAL_NUMBER"),
                "transmittal_date": _parse_date(metadata.get("TRANSMITTAL_DATE")),
                "tag_name": None,
                "property_name": "Not Applicable",
                "group_comment": comment_text,
                "response_vendor": response_text,
                "source_file": metadata.get("SOURCE_FILE"),
                "detail_file": None,
                "detail_sheet": None,
                "comment": "No matching detail sheet found",
                "crs_file_path": str(main_path),
                "crs_file_timestamp": datetime.fromtimestamp(main_path.stat().st_mtime),
            }
            records.append(rec)

    return records


# ---------------------------------------------------------------------------
# Database operations
# ---------------------------------------------------------------------------

def prepare_crs_records(raw_records: list[dict]) -> list[dict]:
    """
    Convert raw parsed records to DB-ready format.
    - Normalize column names (UPPER → snake_case)
    - Generate row_hash and comment_id
    - Add timestamp and status
    """
    db_records = []
    
    for rec in raw_records:
        # Generate row_hash (exclude timestamps)
        row_hash = _calculate_row_hash(rec)
        
        # Generate comment_id: doc_number + row_hash (first 8 chars)
        comment_id = f"{rec.get('doc_number', 'UNK')}#{row_hash[:8]}"
        
        db_record = {
            "comment_id": comment_id,
            "doc_number": rec.get("doc_number"),
            "revision": rec.get("revision"),
            "return_code": rec.get("return_code"),
            "transmittal_number": rec.get("transmittal_number"),
            "transmittal_date": rec.get("transmittal_date"),
            "tag_name": rec.get("tag_name"),
            "property_name": rec.get("property_name"),
            "group_comment": rec.get("group_comment"),
            "response_vendor": rec.get("response_vendor"),
            "source_file": rec.get("source_file"),
            "detail_file": rec.get("detail_file"),
            "detail_sheet": rec.get("detail_sheet"),
            "comment": rec.get("comment"),
            "crs_file_path": rec.get("crs_file_path"),
            "crs_file_timestamp": rec.get("crs_file_timestamp"),
            "row_hash": row_hash,
            "status": "RECEIVED",
            "sync_status": "SYNCED",
            "object_status": "Active",
            # Raw data preservation
            "_raw_doc_number": rec.get("doc_number"),
            "_raw_tag_name": rec.get("tag_name"),
            "_raw_group_comment": rec.get("group_comment"),
            "_raw_comment": rec.get("comment"),
        }
        db_records.append(db_record)
    
    return db_records


def upsert_crs_comments(engine: Engine, records: list[dict], batch_size: int = 500) -> dict:
    """
    Upsert CRS comments using ON CONFLICT.
    Returns stats: {inserted, updated, errors}
    """
    if not records:
        return {"inserted": 0, "updated": 0, "errors": 0}
    
    stats = {"inserted": 0, "updated": 0, "errors": 0}
    
    # Process in batches
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        
        try:
            with engine.connect() as conn:
                for record in batch:
                    stmt = insert(
                        text("""
                            INSERT INTO audit_core.crs_comment (
                                comment_id, doc_number, revision, return_code, transmittal_number,
                                transmittal_date, tag_name, property_name, group_comment,
                                response_vendor, source_file, detail_file, detail_sheet, comment,
                                crs_file_path, crs_file_timestamp, row_hash, status,
                                sync_status, object_status, _raw_doc_number, _raw_tag_name,
                                _raw_group_comment, _raw_comment
                            ) VALUES (
                                :comment_id, :doc_number, :revision, :return_code, :transmittal_number,
                                :transmittal_date, :tag_name, :property_name, :group_comment,
                                :response_vendor, :source_file, :detail_file, :detail_sheet, :comment,
                                :crs_file_path, :crs_file_timestamp, :row_hash, :status,
                                :sync_status, :object_status, :_raw_doc_number, :_raw_tag_name,
                                :_raw_group_comment, :_raw_comment
                            )
                            ON CONFLICT (comment_id) DO UPDATE SET
                                sync_status = CASE WHEN crs_comment.row_hash != :row_hash
                                                   THEN 'MODIFIED'
                                                   ELSE 'SYNCED' END,
                                sync_timestamp = now()
                        """)
                    )
                    
                    # Note: sqlalchemy insert() syntax differs; using raw SQL instead
                    conn.execute(
                        text("""
                            INSERT INTO audit_core.crs_comment (
                                comment_id, doc_number, revision, return_code, transmittal_number,
                                transmittal_date, tag_name, property_name, group_comment,
                                response_vendor, source_file, detail_file, detail_sheet, comment,
                                crs_file_path, crs_file_timestamp, row_hash, status,
                                sync_status, object_status, _raw_doc_number, _raw_tag_name,
                                _raw_group_comment, _raw_comment
                            ) VALUES (
                                :comment_id, :doc_number, :revision, :return_code, :transmittal_number,
                                :transmittal_date, :tag_name, :property_name, :group_comment,
                                :response_vendor, :source_file, :detail_file, :detail_sheet, :comment,
                                :crs_file_path, :crs_file_timestamp, :row_hash, :status,
                                :sync_status, :object_status, :_raw_doc_number, :_raw_tag_name,
                                :_raw_group_comment, :_raw_comment
                            )
                            ON CONFLICT (comment_id) DO UPDATE SET
                                sync_status = CASE WHEN crs_comment.row_hash != :row_hash
                                                   THEN 'MODIFIED'
                                                   ELSE 'SYNCED' END,
                                sync_timestamp = now()
                        """),
                        record
                    )
                
                conn.commit()
                stats["inserted"] += len(batch)
                
        except Exception as exc:
            log.error(f"Error upserting batch: {exc}")
            stats["errors"] += len(batch)
    
    return stats


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_and_load(db_url: str) -> dict:
    """
    Discover, parse, and load CRS files to database.
    Returns summary stats.
    """
    main_files, detail_files = discover_files(SOURCE_DIR)

    work_items = [
        (key, path, tuple(detail_files.get(key, [])))
        for key, path in main_files.items()
    ]

    log.info(
        "Starting parallel processing — %d worker(s), %d job(s).",
        MAX_WORKERS, len(work_items),
    )

    all_records: list[dict] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(process_key, key, path, details): key
            for key, path, details in work_items
        }
        for future in as_completed(futures):
            key = futures[future]
            try:
                records = future.result()
                all_records.extend(records)
                log.info("  ✓ %s — %d record(s)", key, len(records))
            except Exception as exc:
                log.error("  ✗ %s failed: %s", key, exc)

    log.info("Total records parsed: %d", len(all_records))

    # Prepare DB records (add hashes, IDs, etc.)
    db_records = prepare_crs_records(all_records)

    # Connect and upsert
    engine = create_engine(db_url)
    stats = upsert_crs_comments(engine, db_records, batch_size=BATCH_SIZE)
    engine.dispose()

    log.info("Upsert complete: %d inserted, %d updated, %d errors",
             stats["inserted"], stats["updated"], stats["errors"])

    return {
        "files_processed": len(work_items),
        "records_parsed": len(all_records),
        "records_loaded": stats["inserted"] + stats["updated"],
        "errors": stats["errors"],
    }


if __name__ == "__main__":
    import os
    
    # Get DB URL from environment or use default
    db_url = os.getenv(
        "DB_URL",
        "postgresql://postgres:postgres@localhost:5432/engineering_core"
    )
    
    result = process_and_load(db_url)
    print(f"\n✓ Processing complete: {result}")
