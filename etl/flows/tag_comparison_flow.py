"""Standalone Prefect flow: Tag Comparison Report (two-sheet XLSX)."""

import sys
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from prefect import flow, task, get_run_logger
from prefect.cache_policies import NO_CACHE
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

# Setup import path so tasks.* resolves from etl/
current_dir = Path(__file__).resolve().parent
script_root = current_dir.parent
if str(script_root) not in sys.path:
    sys.path.append(str(script_root))

try:
    from tasks.common import load_config, get_db_engine_url
except ImportError as e:
    print(f"[SKIP] {Path(__file__).name}: Could not import task modules. Details: {e}")
    sys.exit(0)

# Module-level config — same pattern as other export flows
config = load_config()
DB_URL = get_db_engine_url(config)
_EXPORT_DIR = config.get("storage", {}).get("export_dir", ".")

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# All columns loaded from project_core.tag (comparison candidates + metadata)
_TAG_COLUMNS: list[str] = [
    "source_id",
    "tag_name",
    "tag_status",
    "description",
    "parent_tag_raw",
    "tag_class_raw",
    "area_code_raw",
    "process_unit_raw",
    "discipline_code_raw",
    "po_code_raw",
    "design_company_name_raw",
    "company_raw",
    "manufacturer_company_raw",
    "vendor_company_raw",
    "article_code_raw",
    "model_part_raw",
    "safety_critical_item",
    "safety_critical_item_reason_awarded",
    "production_critical_item",
    "serial_no",
    "install_date",
    "startup_date",
    "warranty_end_date",
    "price",
    "tech_id",
    "ip_grade",
    "ex_class",
    "mc_package_code",
    "equip_no",
    "alias",
    "from_tag_raw",
    "to_tag_raw",
    "plant_raw",
    "sync_status",
    "object_status",
]

# Columns excluded from field-level comparison (anchor, hashes, metadata)
_EXCLUDE_FROM_COMPARISON: frozenset[str] = frozenset(
    {"source_id", "id", "row_hash", "sync_timestamp", "run_id", "sync_status", "object_status"}
)

# Fields used for actual field-level diffing
_DATA_COLS: list[str] = [c for c in _TAG_COLUMNS if c not in _EXCLUDE_FROM_COMPARISON]

# Date columns that need normalisation to YYYY-MM-DD for consistent comparison
_DATE_COLS: frozenset[str] = frozenset({"install_date", "startup_date", "warranty_end_date"})

# ---------------------------------------------------------------------------
# SQL queries
# ---------------------------------------------------------------------------

_SQL_CURRENT = """
/*
Purpose: Load current tag state for comparison report.
Gate:    object_status = 'Active', sync_timestamp <= current_date.
Changes: 2026-04-16 — Initial implementation.
*/
SELECT
    source_id,
    tag_name,
    tag_status,
    description,
    parent_tag_raw,
    tag_class_raw,
    area_code_raw,
    process_unit_raw,
    discipline_code_raw,
    po_code_raw,
    design_company_name_raw,
    company_raw,
    manufacturer_company_raw,
    vendor_company_raw,
    article_code_raw,
    model_part_raw,
    safety_critical_item,
    safety_critical_item_reason_awarded,
    production_critical_item,
    serial_no,
    install_date,
    startup_date,
    warranty_end_date,
    price,
    tech_id,
    ip_grade,
    ex_class,
    mc_package_code,
    equip_no,
    alias,
    from_tag_raw,
    to_tag_raw,
    plant_raw,
    sync_status,
    object_status,
    row_hash
FROM project_core.tag
WHERE object_status = 'Active'
  AND sync_timestamp::date <= :current_date
"""

_SQL_BASELINE = """
/*
Purpose: Most recent per-tag snapshot on or before baseline_date.
Gate:    sync_timestamp::date <= baseline_date.
Note:    DISTINCT ON picks the latest history record per tag.
Changes: 2026-04-16 — Initial implementation.
*/
SELECT DISTINCT ON (source_id)
    source_id,
    tag_name,
    row_hash,
    snapshot,
    sync_timestamp
FROM audit_core.tag_status_history
WHERE sync_timestamp::date <= :baseline_date
ORDER BY source_id, sync_timestamp DESC
"""

_SQL_MIN_HISTORY_DATE = """
SELECT MIN(sync_timestamp)::date AS min_date
FROM audit_core.tag_status_history
"""

# ---------------------------------------------------------------------------
# Style constants (openpyxl)
# ---------------------------------------------------------------------------

_FILL_GREEN_HEADER = PatternFill(fill_type="solid", fgColor="C6EFCE")   # _new headers
_FILL_BLUE_HEADER = PatternFill(fill_type="solid", fgColor="BDD7EE")    # _old headers
_FILL_GREY = PatternFill(fill_type="solid", fgColor="D9D9D9")           # neutral headers
_FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")         # Modified cell
_FILL_ROW_GREEN = PatternFill(fill_type="solid", fgColor="E2EFDA")      # Created row
_FILL_ROW_RED = PatternFill(fill_type="solid", fgColor="FCE4D6")        # Deleted row

_FONT_GREEN = Font(name="Calibri", size=11, bold=True, color="276221")
_FONT_BLUE = Font(name="Calibri", size=11, bold=True, color="1F497D")
_FONT_BLACK_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
_FONT_DEFAULT = Font(name="Calibri", size=11)

_CENTER = Alignment(horizontal="center")

# Sort priority for Comparison_Result
_RESULT_SORT_ORDER: dict[str, int] = {
    "Modified": 0,
    "Created": 1,
    "Deleted": 2,
    "No Changes": 3,
}

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def _to_str(value) -> str:
    """Convert any value to string; None/NaN/NaT map to empty string."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def _escape_formula(value: str) -> str:
    """Prepend a space if value starts with '=' to prevent formula injection."""
    if value.startswith("="):
        return " " + value
    return value


def _safe_cell(value) -> str:
    """Convert value to safe cell string (None→'', formula-escaped)."""
    return _escape_formula(_to_str(value))


def _normalise_date_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise date columns to YYYY-MM-DD string for consistent comparison."""
    for col in _DATE_COLS:
        if col in df.columns:
            df[col] = (
                pd.to_datetime(df[col], errors="coerce")
                .dt.strftime("%Y-%m-%d")
                .fillna("")
            )
    return df


def _autofit_columns(
    ws, min_width: int = 12, max_width: int = 50, padding: int = 2
) -> None:
    """Set each column width based on the longest cell value in that column."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(
            max_width, max(min_width, max_len + padding)
        )


# ---------------------------------------------------------------------------
# Comparison logic
# ---------------------------------------------------------------------------


def _get_comparison_result(row: pd.Series, data_cols: list[str]) -> str:
    """
    Classify a merged row as Modified, Created, Deleted, or No Changes.

    Mirrors ReportComparisionTool.py logic exactly:
    - Fast path: identical row_hash → No Changes (avoids field scan)
    - VOID tag_status → Deleted
    - left_only (not in baseline) → Created
    - right_only (not in current) → Deleted

    Args:
        row: Series from the outer-merged DataFrame.
        data_cols: Column names used for field-level comparison.

    Returns:
        One of 'Modified', 'Created', 'Deleted', 'No Changes'.
    """
    merge_flag = row["_merge"]
    tag_status_new = _to_str(row.get("tag_status_new", ""))

    if merge_flag == "both":
        # Fast path: skip field scan when hash is identical
        if row.get("row_hash_new") == row.get("row_hash_old"):
            return "No Changes"
        if tag_status_new == "VOID":
            return "Deleted"
        changed = any(
            _to_str(row.get(f"{c}_new", "")) != _to_str(row.get(f"{c}_old", ""))
            for c in data_cols
        )
        return "Modified" if changed else "No Changes"

    if merge_flag == "left_only":
        return "Deleted" if tag_status_new == "VOID" else "Created"

    if merge_flag == "right_only":
        return "Deleted"

    return ""


# ---------------------------------------------------------------------------
# Prefect tasks
# ---------------------------------------------------------------------------


@task(name="load-current-tags", retries=1, cache_policy=NO_CACHE)
def load_current_tags(engine: Engine, current_date: date) -> pd.DataFrame:
    """
    Load all active tags as of current_date from project_core.tag.

    All values are coerced to str; date columns normalised to YYYY-MM-DD.

    Args:
        engine: SQLAlchemy engine connected to engineering_core.
        current_date: Upper bound (inclusive) for sync_timestamp.

    Returns:
        DataFrame keyed on source_id with all comparison columns.
    """
    logger = get_run_logger()
    with engine.connect() as conn:
        df = pd.read_sql(
            text(_SQL_CURRENT).bindparams(current_date=str(current_date)),
            conn,
        )
    df = _normalise_date_cols(df)
    # Coerce all columns (except the merge key) to str
    for col in df.columns:
        if col != "source_id":
            df[col] = df[col].apply(_to_str)
    logger.info(f"Loaded {len(df)} current tags as of {current_date}")
    return df


@task(name="load-baseline-tags", retries=1, cache_policy=NO_CACHE)
def load_baseline_tags(engine: Engine, baseline_date: date) -> pd.DataFrame:
    """
    Load the most recent snapshot per tag recorded on or before baseline_date.

    Expands the snapshot JSONB into columns matching project_core.tag field
    names. Missing snapshot keys default to ''.

    Args:
        engine: SQLAlchemy engine connected to engineering_core.
        baseline_date: Upper bound (inclusive) for sync_timestamp.

    Returns:
        DataFrame keyed on source_id with row_hash and all data_cols.

    Raises:
        ValueError: If no history records exist on or before baseline_date.
    """
    logger = get_run_logger()
    with engine.connect() as conn:
        df = pd.read_sql(
            text(_SQL_BASELINE).bindparams(baseline_date=str(baseline_date)),
            conn,
        )

    if df.empty:
        with engine.connect() as conn:
            row = conn.execute(text(_SQL_MIN_HISTORY_DATE)).fetchone()
        min_ts = row[0] if row else "unknown"
        raise ValueError(
            f"No history records found on or before {baseline_date}. "
            f"Earliest available: {min_ts}"
        )

    # Expand snapshot JSONB → one column per key
    snapshot_expanded: pd.DataFrame = df["snapshot"].apply(
        lambda s: s if isinstance(s, dict) else {}
    ).apply(pd.Series)

    # Ensure all data_cols are present; fill any missing ones with ''
    for col in _DATA_COLS:
        if col not in snapshot_expanded.columns:
            snapshot_expanded[col] = ""

    snapshot_expanded = snapshot_expanded.reindex(columns=_DATA_COLS, fill_value="")

    # Normalise date columns from JSON strings for consistent comparison
    snapshot_expanded = _normalise_date_cols(snapshot_expanded)

    # Coerce everything to str
    for col in snapshot_expanded.columns:
        snapshot_expanded[col] = snapshot_expanded[col].apply(_to_str)

    # Combine anchor columns with expanded snapshot
    result_df = pd.concat(
        [
            df[["source_id", "row_hash"]].reset_index(drop=True),
            snapshot_expanded.reset_index(drop=True),
        ],
        axis=1,
    )
    logger.info(f"Loaded {len(result_df)} baseline snapshots as of {baseline_date}")
    return result_df


# ---------------------------------------------------------------------------
# XLSX sheet builders
# ---------------------------------------------------------------------------


def _build_full_comparison_sheet(
    ws, merged: pd.DataFrame, data_cols: list[str]
) -> None:
    """
    Write Sheet 1 "Full Comparison".

    Column layout: [_new block] | source_id | [_old block] | Comparison_Result
    Colour rules:
      - _new header: light green / dark green bold
      - _old header: light blue / dark blue bold
      - source_id / Comparison_Result header: grey / black bold
      - Modified cell: yellow on cells where field_new != field_old
      - Created row: light green on all _new cells
      - Deleted row: light red on all _old cells
    Sort order: Modified → Created → Deleted → No Changes, then source_id ASC.

    Args:
        ws: openpyxl Worksheet to write into.
        merged: Outer-merged DataFrame with Comparison_Result column.
        data_cols: Ordered list of field names used for comparison.
    """
    new_headers = [f"{c}_new" for c in data_cols]
    old_headers = [f"{c}_old" for c in data_cols]
    all_headers = new_headers + ["source_id"] + old_headers + ["Comparison_Result"]

    # --- Header row ---
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = _CENTER
        if header.endswith("_new"):
            cell.fill = _FILL_GREEN_HEADER
            cell.font = _FONT_GREEN
        elif header.endswith("_old"):
            cell.fill = _FILL_BLUE_HEADER
            cell.font = _FONT_BLUE
        else:
            cell.fill = _FILL_GREY
            cell.font = _FONT_BLACK_BOLD

    # --- Sort ---
    df_sorted = merged.copy()
    df_sorted["_sort_key"] = df_sorted["Comparison_Result"].map(
        lambda r: _RESULT_SORT_ORDER.get(r, 99)
    )
    df_sorted.sort_values(
        ["_sort_key", "source_id"], na_position="last", inplace=True
    )
    df_sorted.reset_index(drop=True, inplace=True)

    # --- Data rows ---
    for excel_row, (_, row) in enumerate(df_sorted.iterrows(), start=2):
        result = _to_str(row.get("Comparison_Result", ""))

        for col_idx, header in enumerate(all_headers, start=1):
            cell_val = _safe_cell(row.get(header, ""))
            cell = ws.cell(row=excel_row, column=col_idx, value=cell_val)
            cell.font = _FONT_DEFAULT

            # Cell-level fill
            if result == "Modified":
                # Highlight only cells where the field actually changed
                if header.endswith("_new") or header.endswith("_old"):
                    base = header[:-4]  # strip '_new' or '_old'
                    val_new = _to_str(row.get(f"{base}_new", ""))
                    val_old = _to_str(row.get(f"{base}_old", ""))
                    if val_new != val_old:
                        cell.fill = _FILL_YELLOW
            elif result == "Created" and header.endswith("_new"):
                cell.fill = _FILL_ROW_GREEN
            elif result == "Deleted" and header.endswith("_old"):
                cell.fill = _FILL_ROW_RED

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws)


def _build_changes_only_sheet(
    ws, merged: pd.DataFrame, data_cols: list[str]
) -> None:
    """
    Write Sheet 2 "Changes Only".

    One row per changed field for Modified/Created/Deleted tags:
      - Modified: only fields where field_new != field_old
      - Created:  all fields; VALUE_OLD = ''
      - Deleted:  all fields; VALUE_NEW = ''

    Columns: SOURCE_ID | TAG_NAME | FIELD_NAME | VALUE_OLD | VALUE_NEW | Comparison_Result

    Args:
        ws: openpyxl Worksheet to write into.
        merged: Outer-merged DataFrame with Comparison_Result column.
        data_cols: Ordered list of field names used for comparison.
    """
    headers = [
        "SOURCE_ID",
        "TAG_NAME",
        "FIELD_NAME",
        "VALUE_OLD",
        "VALUE_NEW",
        "Comparison_Result",
    ]
    header_fills = {
        "SOURCE_ID": _FILL_GREY,
        "TAG_NAME": _FILL_GREY,
        "FIELD_NAME": _FILL_GREY,
        "VALUE_OLD": _FILL_BLUE_HEADER,
        "VALUE_NEW": _FILL_GREEN_HEADER,
        "Comparison_Result": _FILL_GREY,
    }
    header_fonts = {
        "SOURCE_ID": _FONT_BLACK_BOLD,
        "TAG_NAME": _FONT_BLACK_BOLD,
        "FIELD_NAME": _FONT_BLACK_BOLD,
        "VALUE_OLD": _FONT_BLUE,
        "VALUE_NEW": _FONT_GREEN,
        "Comparison_Result": _FONT_BLACK_BOLD,
    }

    # --- Header row ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fills[header]
        cell.font = header_fonts[header]
        cell.alignment = _CENTER

    # --- Collect changed rows ---
    changed = merged[
        merged["Comparison_Result"].isin(["Modified", "Created", "Deleted"])
    ]

    sheet2_rows: list[tuple[str, str, str, str, str, str]] = []

    for _, row in changed.iterrows():
        source_id = _to_str(row.get("source_id", ""))
        result = _to_str(row.get("Comparison_Result", ""))
        tag_name = _to_str(row.get("tag_name_new", "")) or _to_str(
            row.get("tag_name_old", "")
        )

        if result == "Modified":
            for field in data_cols:
                val_new = _to_str(row.get(f"{field}_new", ""))
                val_old = _to_str(row.get(f"{field}_old", ""))
                if val_new != val_old:
                    sheet2_rows.append(
                        (source_id, tag_name, field, val_old, val_new, result)
                    )

        elif result == "Created":
            for field in data_cols:
                val_new = _to_str(row.get(f"{field}_new", ""))
                sheet2_rows.append((source_id, tag_name, field, "", val_new, result))

        elif result == "Deleted":
            for field in data_cols:
                val_old = _to_str(row.get(f"{field}_old", ""))
                sheet2_rows.append((source_id, tag_name, field, val_old, "", result))

    # Sort: SOURCE_ID ASC, FIELD_NAME ASC
    sheet2_rows.sort(key=lambda r: (r[0], r[2]))

    _row_fill: dict[str, PatternFill] = {
        "Modified": _FILL_YELLOW,
        "Created": _FILL_ROW_GREEN,
        "Deleted": _FILL_ROW_RED,
    }

    # --- Data rows ---
    for excel_row, (src_id, tag_nm, field, val_old, val_new, result) in enumerate(
        sheet2_rows, start=2
    ):
        values = [
            src_id,
            tag_nm,
            field,
            _escape_formula(val_old),
            _escape_formula(val_new),
            result,
        ]
        fill = _row_fill.get(result)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = _FONT_DEFAULT
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Prefect flow
# ---------------------------------------------------------------------------


@flow(name="export_tag_comparison", log_prints=True)
def export_tag_comparison_flow(
    current_date: Optional[date] = None,
    baseline_date: Optional[date] = None,
    output_dir: Optional[str] = None,
) -> str:
    """
    Generates a two-sheet XLSX tag comparison report.

    Args:
        current_date:  Reference date for "new" state.
                       Default: today.
        baseline_date: Reference date for "old" (baseline) state.
                       Default: current_date minus 1 calendar month
                       (dateutil.relativedelta).
        output_dir:    Override output directory.
                       Default: same output root as other export flows.

    Returns:
        Absolute path of the written XLSX file.

    Output filename:
        tag-comparison-{current_date_str}-vs-{baseline_date_str}.xlsx
        Example: tag-comparison-2026-04-16-vs-2026-03-16.xlsx

    Example:
        >>> export_tag_comparison_flow(
        ...     current_date=date(2026, 4, 16),
        ...     baseline_date=date(2026, 3, 16),
        ... )
        '/mnt/shared-data/.../tag-comparison-2026-04-16-vs-2026-03-16.xlsx'
    """
    logger = get_run_logger()

    # Resolve date defaults
    if current_date is None:
        current_date = date.today()
    if baseline_date is None:
        # One calendar month back (handles month-end edge cases correctly)
        baseline_date = current_date - relativedelta(months=1)

    current_date_str = current_date.strftime("%Y-%m-%d")
    baseline_date_str = baseline_date.strftime("%Y-%m-%d")
    filename = f"tag-comparison-{current_date_str}-vs-{baseline_date_str}.xlsx"
    output_path = Path(output_dir or _EXPORT_DIR) / filename

    engine = create_engine(DB_URL)

    # Load both states
    df_new = load_current_tags(engine, current_date)
    df_old = load_baseline_tags(engine, baseline_date)

    logger.info(f"Baseline date: {baseline_date} | Current date: {current_date}")
    logger.info(
        f"Tags in current: {len(df_new)} | Tags in baseline: {len(df_old)}"
    )

    # Trim to join key + row_hash + data_cols only
    new_keep = ["source_id", "row_hash"] + [
        c for c in _DATA_COLS if c in df_new.columns
    ]
    old_keep = ["source_id", "row_hash"] + [
        c for c in _DATA_COLS if c in df_old.columns
    ]
    df_new_trim = df_new[new_keep].copy()
    df_old_trim = df_old[old_keep].copy()

    # Outer merge on the immutable source system identifier
    merged = pd.merge(
        df_new_trim,
        df_old_trim,
        on="source_id",
        suffixes=("_new", "_old"),
        how="outer",
        indicator=True,
    )

    # Classify every row
    merged["Comparison_Result"] = merged.apply(
        lambda row: _get_comparison_result(row, _DATA_COLS),
        axis=1,
    )

    n_mod = int((merged["Comparison_Result"] == "Modified").sum())
    n_cre = int((merged["Comparison_Result"] == "Created").sum())
    n_del = int((merged["Comparison_Result"] == "Deleted").sum())
    n_nc = int((merged["Comparison_Result"] == "No Changes").sum())

    logger.info(
        f"Modified: {n_mod} | Created: {n_cre} | Deleted: {n_del} | No Changes: {n_nc}"
    )

    # Build workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Full Comparison"
    _build_full_comparison_sheet(ws1, merged, _DATA_COLS)

    ws2 = wb.create_sheet(title="Changes Only")
    _build_changes_only_sheet(ws2, merged, _DATA_COLS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    logger.info(f"Output: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    _REPO_ROOT = Path(__file__).resolve().parent.parent.parent
    export_tag_comparison_flow.from_source(
        source=str(_REPO_ROOT),
        entrypoint="etl/flows/tag_comparison_flow.py:export_tag_comparison_flow",
    ).deploy(
        name="export_tag_comparison_deploy",
        work_pool_name="default-agent-pool",
    )
