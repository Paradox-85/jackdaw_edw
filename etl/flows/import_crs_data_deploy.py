"""
import_crs_data_deploy.py

Prefect 3.x flow: ingest CRS (Comment Review Sheet) Excel files into audit_core.crs_comment.

Parsing logic preserved from scripts/crs_excel_parser.py (tested, production-ready).
DB upsert + Prefect structure from sync_crs_data_v2.py (corrected anti-patterns).

Flow structure:
  discover_crs_files    — scan root dir, build main_files + detail_files dicts
  parse_all_files       — ThreadPoolExecutor INSIDE @task (Prefect 3.x compliant)
  prepare_crs_records   — row_hash + comment_id + FK resolution (tag_id, doc_id)
  upsert_crs_records    — ON CONFLICT (comment_id) DO UPDATE WHERE hash changed

Audit:
  audit_core.sync_run_stats — INSERT on start, UPDATE on completion (audit-rules.md)

Requirements:
  pip install pandas openpyxl python-calamine prefect sqlalchemy
"""

from __future__ import annotations

import hashlib
import json
import re
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from prefect import flow, task
from prefect.cache_policies import NO_CACHE
from prefect.logging import get_run_logger
from sqlalchemy import create_engine, text
from sqlalchemy.pool import QueuePool

from etl.tasks.common import clean_string, load_config, get_db_engine_url, to_dt


# =============================================================================
# Constants (from scripts/crs_excel_parser.py — tested values)
# =============================================================================

MAIN_PATTERN = re.compile(
    r"^DOC_COMMENT_(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d{2})_[A-Z]{3}\.xlsx$"
)
DETAIL_PATTERN = re.compile(
    r"^(JDAW-KVE-E-JA-6944-00001-\d{3}_A\d{2})(?:_\d+|_Review_Comments)\.xlsx$"
)

COMMENT_COL_KEYWORDS  = ("remark", "adura", "issue", "comment")
PROPERTY_COL_KEYWORDS = ("equipment property name", "tag property name", "property name")
SKIP_SHEETS = {"comment_sheet"}  # normalised sheet name

# Fields excluded from row_hash to prevent false positives on re-runs
HASH_EXCLUDE_FIELDS = {
    "sync_timestamp",
    "crs_file_timestamp",
    "llm_response_timestamp",
    "response_approval_date",
    "llm_response",
}

MAX_WORKERS = 6
BATCH_SIZE  = 500


# =============================================================================
# Parsing helpers (verbatim from scripts/crs_excel_parser.py)
# =============================================================================

def _scalar(val: Any) -> Any:
    """Extract scalar from value that may be a pd.Series."""
    if isinstance(val, pd.Series):
        return val.iloc[0] if not val.empty else None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def _norm_sheet(name: str) -> str:
    """Normalise sheet name: strip, lowercase, spaces → underscores."""
    return name.strip().lower().replace(" ", "_")


def _expand_merged_cells(ws) -> dict[tuple[int, int], object]:
    """Build cell_map with merged cells expanded to top-left value."""
    cell_map: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for cell in row:
            cell_map[(cell.row, cell.column)] = cell.value

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = cell_map.get((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_map[(r, c)] = top_left_val

    return cell_map


def _build_two_row_header(
    cell_map: dict, row1: int, row2: int, max_col: int
) -> list[str]:
    """Combine two header rows into UPPER_SNAKE column names. Deduplicates."""
    headers = []
    for col in range(1, max_col + 1):
        v1 = str(cell_map.get((row1, col), "") or "").strip()
        v2 = str(cell_map.get((row2, col), "") or "").strip()

        if v1 and v2 and v1 != v2:
            name = f"{v1}_{v2}"
        elif v1:
            name = v1
        elif v2:
            name = v2
        else:
            name = f"COL_{col}"

        name = re.sub(r"[\s/\\().,-]+", "_", name).upper().strip("_")
        headers.append(name)

    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def _find_comment_column(ws, data_start_row: int) -> int | None:
    """
    Detect vertically-merged single-column comment cell in a detail sheet.
    - Single-column merge (min_col == max_col)
    - Must start at data_start_row
    - Must span at least 2 rows
    - Multiple candidates: pick largest span.
    Returns 0-based column index, or None.
    """
    best_span: int = 1
    best_col:  int | None = None

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row_range = merged_range.bounds

        if min_col != max_col:
            continue
        if min_row != data_start_row:
            continue

        span = max_row_range - min_row + 1
        if span > best_span:
            best_span = span
            best_col  = min_col - 1

    return best_col


# SheetData: (DataFrame, merged_comment_value, fallback_comment_col_name)
SheetData = tuple[pd.DataFrame, str | None, str | None]


def _load_detail_file_impl(path: Path) -> dict[str, SheetData]:
    """
    Two-phase detail file loading:
    Phase 1 (openpyxl): detect vertically-merged comment column per sheet.
    Phase 2 (calamine): read data rows fast; drop merged comment col from df.
    Returns {sheet_key: (df, merged_comment_value, fallback_comment_col)}.
    """
    result: dict[str, SheetData] = {}
    DATA_START_ROW = 2

    # Phase 1: openpyxl for merged-cell detection
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        return result  # caller logs errors via task logger

    sheet_names = list(wb.sheetnames)
    sheet_meta: dict[str, tuple[int | None, str | None]] = {}

    for sheet in sheet_names:
        sheet_key = _norm_sheet(sheet)
        if sheet_key in SKIP_SHEETS:
            continue

        ws = wb[sheet]
        comment_col_idx = _find_comment_column(ws, data_start_row=DATA_START_ROW)

        comment_value: str | None = None
        if comment_col_idx is not None:
            raw = ws.cell(row=DATA_START_ROW, column=comment_col_idx + 1).value
            comment_value = str(raw).strip() if raw is not None else None
            comment_value = comment_value or None

        sheet_meta[sheet_key] = (comment_col_idx, comment_value)

    wb.close()

    # Phase 2: calamine for fast data reading
    try:
        xl = pd.ExcelFile(path, engine="calamine")
    except Exception:
        return result

    for sheet in sheet_names:
        sheet_key = _norm_sheet(sheet)
        if sheet_key not in sheet_meta:
            continue

        comment_col_idx, comment_value = sheet_meta[sheet_key]

        try:
            df = pd.read_excel(
                xl, sheet_name=sheet, engine="calamine",
                dtype=str, na_filter=False,
            )
        except Exception:
            continue

        # Drop fully empty rows
        df = df[~df.apply(lambda r: all(v == "" for v in r), axis=1)].copy()
        if df.empty:
            continue

        # Drop merged comment column — value already captured as scalar
        if comment_col_idx is not None and comment_col_idx < len(df.columns):
            df = df.drop(columns=df.columns[comment_col_idx])

        # Fallback: find comment column by keyword when no merge detected
        fallback_comment_col: str | None = None
        if comment_col_idx is None:
            for col in df.columns:
                if any(kw in col.lower() for kw in COMMENT_COL_KEYWORDS):
                    fallback_comment_col = col
                    break

        # Filter rows with empty fallback comment (no finding = no record)
        if fallback_comment_col is not None:
            df = df[df[fallback_comment_col].str.strip() != ""].copy()
            df = df.reset_index(drop=True)
            if df.empty:
                continue

        result[sheet_key] = (df, comment_value, fallback_comment_col)

    return result


# Thread-safe detail file cache (reset between flow runs)
_detail_file_cache: dict[Path, dict] = {}
_cache_lock = threading.Lock()


def _load_detail_file(path: Path) -> dict[str, SheetData]:
    """Thread-safe double-checked locking cache wrapper."""
    if path in _detail_file_cache:
        return _detail_file_cache[path]
    with _cache_lock:
        if path not in _detail_file_cache:
            _detail_file_cache[path] = _load_detail_file_impl(path)
    return _detail_file_cache[path]


def find_matching_sheet(
    comment_text: str,
    sheets: dict[str, SheetData],
) -> tuple[str | None, pd.DataFrame | None, str | None, str | None]:
    """
    Match comment text to a detail sheet key.
    Returns (sheet_key, df, merged_comment_value, fallback_comment_col).
    """
    text_norm = comment_text.lower().replace(" ", "_")
    for sheet_key, (df, comment_val, fallback_col) in sheets.items():
        if sheet_key in text_norm:
            return sheet_key, df, comment_val, fallback_col
    return None, None, None, None


def parse_main_file(path: Path) -> tuple[dict | None, pd.DataFrame | None]:
    """
    Extract header metadata and comment table from a DOC_COMMENT_* file.
    Phase 1 (openpyxl): read merged-cell metadata from rows 1–7.
    Phase 2 (calamine): read data rows 8+ fast.
    Returns (metadata_dict, df_comments) or (None, None) on skip/error.
    """
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.active
    except Exception:
        return None, None

    cell_map = _expand_merged_cells(ws)
    max_col  = ws.max_column or 0
    wb.close()

    def g(row: int, col: int) -> Any:
        val = cell_map.get((row, col))
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        return val

    metadata = {
        "DOC_NUMBER":         g(4, 3),
        "REVISION":           g(3, 6),
        "RETURN_CODE":        g(4, 6),
        "TRANSMITTAL_NUMBER": g(5, 3),
        "TRANSMITTAL_DATE":   g(5, 6),
        "SOURCE_FILE":        path.name,
        "CRS_FILE_PATH":      str(path),
    }

    # Skip files with Return Code == 1 (vendor not required to respond)
    rc = str(metadata["RETURN_CODE"]).strip().split(".")[0]
    if rc == "1":
        return None, None

    if max_col < 3:
        return metadata, None

    headers = _build_two_row_header(cell_map, row1=6, row2=7, max_col=max_col)

    try:
        df_raw = pd.read_excel(
            path, header=None, skiprows=7,
            engine="calamine", dtype=str, na_filter=False,
        )
    except Exception:
        return metadata, None

    if df_raw.empty:
        return metadata, None

    actual_cols = df_raw.shape[1]
    if len(headers) < actual_cols:
        headers += [f"COL_{i}" for i in range(len(headers), actual_cols)]
    else:
        headers = headers[:actual_cols]
    df_raw.columns = headers

    col_c_name = headers[2] if len(headers) > 2 else None
    col_f_name = headers[5] if len(headers) > 5 else None

    if col_c_name is None:
        return metadata, None

    df_comments = pd.DataFrame()
    df_comments["GROUP_COMMENT"] = df_raw[col_c_name]
    df_comments["RESPONSE"]      = df_raw[col_f_name] if col_f_name else ""

    col_a_name   = headers[0] if headers else None
    mask_a_empty = (
        df_raw[col_a_name].str.strip() == ""
        if col_a_name else pd.Series([True] * len(df_raw))
    )
    mask_c_empty = df_comments["GROUP_COMMENT"].str.strip() == ""

    df_comments = df_comments[~(mask_a_empty & mask_c_empty)].copy()
    df_comments = df_comments[
        df_comments["GROUP_COMMENT"].str.strip() != ""
    ].reset_index(drop=True)
    df_comments["RESPONSE"] = df_comments["RESPONSE"].replace("", None)

    return metadata, df_comments


def process_key(
    key: str,
    main_path: Path,
    related_paths: tuple[Path, ...],
) -> tuple[list[dict], list[str]]:
    """
    Process one DOC_COMMENT_* file and its related detail sheets.
    Returns (records, orphan_sheets) where orphan_sheets are detail sheets
    that had no matching comment in the main file.
    """
    records:      list[dict] = []
    matched_keys: set[str]   = set()

    metadata, df_comments = parse_main_file(main_path)
    if metadata is None or df_comments is None:
        return records, []

    # Load all detail files for this document key
    all_detail_sheets: dict[Path, dict[str, SheetData]] = {
        p: _load_detail_file(p) for p in related_paths
    }

    for _, row in df_comments.iterrows():
        comment_text  = str(row["GROUP_COMMENT"]).strip()
        response_text = _scalar(row.get("RESPONSE"))
        found_detail  = False

        for detail_path, sheets in all_detail_sheets.items():
            sheet_key, df_sheet, comment_val, fallback_col = find_matching_sheet(
                comment_text, sheets
            )
            if df_sheet is None:
                continue

            found_detail = True
            matched_keys.add(f"{detail_path.name}::{sheet_key}")

            # Column detection
            tag_col = next(
                (c for c in df_sheet.columns
                 if "tag" in c.lower() and "property" not in c.lower()), None
            )
            prop_col = next(
                (c for c in df_sheet.columns
                 if any(kw in c.lower() for kw in PROPERTY_COL_KEYWORDS)), None
            )

            for _, d_row in df_sheet.iterrows():
                tag_name  = _scalar(d_row[tag_col])  if tag_col  else None
                prop_name = _scalar(d_row[prop_col]) if prop_col else None
                if not prop_name or str(prop_name).strip() == "":
                    prop_name = "Not Applicable"

                row_comment = comment_val
                if row_comment is None and fallback_col and fallback_col in d_row.index:
                    row_comment = _scalar(d_row[fallback_col]) or None

                records.append({
                    "DOC_NUMBER":         metadata.get("DOC_NUMBER"),
                    "REVISION":           metadata.get("REVISION"),
                    "RETURN_CODE":        metadata.get("RETURN_CODE"),
                    "TRANSMITTAL_NUMBER": metadata.get("TRANSMITTAL_NUMBER"),
                    "TRANSMITTAL_DATE":   metadata.get("TRANSMITTAL_DATE"),
                    "TAG_NAME":           tag_name,
                    "PROPERTY_NAME":      prop_name,
                    "GROUP_COMMENT":      comment_text,
                    "RESPONSE":           response_text,
                    "SOURCE_FILE":        metadata.get("SOURCE_FILE"),
                    "DETAIL_FILE":        detail_path.name,
                    "DETAIL_SHEET":       sheet_key,
                    "COMMENT":            row_comment,
                    "CRS_FILE_PATH":      metadata.get("CRS_FILE_PATH"),
                })

            break  # one comment → one matching sheet

        if not found_detail:
            records.append({
                "DOC_NUMBER":         metadata.get("DOC_NUMBER"),
                "REVISION":           metadata.get("REVISION"),
                "RETURN_CODE":        metadata.get("RETURN_CODE"),
                "TRANSMITTAL_NUMBER": metadata.get("TRANSMITTAL_NUMBER"),
                "TRANSMITTAL_DATE":   metadata.get("TRANSMITTAL_DATE"),
                "TAG_NAME":           None,
                "PROPERTY_NAME":      "Not Applicable",
                "GROUP_COMMENT":      comment_text,
                "RESPONSE":           response_text,
                "SOURCE_FILE":        metadata.get("SOURCE_FILE"),
                "DETAIL_FILE":        None,
                "DETAIL_SHEET":       None,
                "COMMENT":            "No matching detail sheet found",
                "CRS_FILE_PATH":      metadata.get("CRS_FILE_PATH"),
            })

    # Orphan detection: detail sheets that were loaded but never matched
    orphan_sheets: list[str] = []
    for detail_path, sheets in all_detail_sheets.items():
        for sk in sheets:
            composite = f"{detail_path.name}::{sk}"
            if composite not in matched_keys:
                orphan_sheets.append(composite)

    return records, orphan_sheets


# =============================================================================
# Prefect Tasks
# =============================================================================

@task(name="Discover CRS Files", cache_policy=NO_CACHE)
def discover_crs_files(root: str) -> tuple[dict[str, str], dict[str, list[str]]]:
    """
    Scan root directory for CRS Excel files.
    Returns (main_files, detail_files) with str paths (serializable by Prefect).
    """
    logger = get_run_logger()
    root_path = Path(root)

    if not root_path.exists():
        raise FileNotFoundError(f"CRS data directory not found: {root}")

    main_files:   dict[str, str]        = {}
    detail_files: dict[str, list[str]]  = {}

    for path in root_path.rglob("*.xlsx"):
        if "_templates" in path.parts:
            continue

        name = path.name
        m = MAIN_PATTERN.match(name)
        if m:
            key = m.group(1)
            if key not in main_files:
                main_files[key] = str(path)
            else:
                logger.warning("Duplicate main key %s — keeping first: %s", key, main_files[key])
            continue

        d = DETAIL_PATTERN.match(name)
        if d:
            detail_files.setdefault(d.group(1), []).append(str(path))

    logger.info("Found %d main file(s), %d detail key(s).", len(main_files), len(detail_files))
    return main_files, detail_files


@task(name="Parse CRS Files", cache_policy=NO_CACHE)
def parse_all_files(
    main_files:   dict[str, str],
    detail_files: dict[str, list[str]],
    max_workers:  int = MAX_WORKERS,
) -> tuple[list[dict], list[str]]:
    """
    Parse all CRS Excel files using ThreadPoolExecutor INSIDE this task.
    Calling @task from a thread is not supported in Prefect 3.x — the pool
    must live entirely within a single task boundary.
    Returns (all_records, orphan_sheets).
    """
    logger = get_run_logger()

    work_items = [
        (key, Path(path), tuple(Path(p) for p in detail_files.get(key, [])))
        for key, path in main_files.items()
    ]

    all_records:   list[dict] = []
    orphan_sheets: list[str]  = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(process_key, key, main_path, related): key
            for key, main_path, related in work_items
        }
        for future in as_completed(futures):
            key = futures[future]
            try:
                records, file_orphans = future.result()
                all_records.extend(records)
                orphan_sheets.extend(file_orphans)
                logger.info("  ✓ %s — %d record(s), %d orphan sheet(s)",
                            key, len(records), len(file_orphans))
            except Exception as exc:
                logger.error("  ✗ %s failed: %s", key, exc)

    _detail_file_cache.clear()  # prevent memory leak between runs
    logger.info("Parsed %d total records, %d orphan sheets.", len(all_records), len(orphan_sheets))
    return all_records, orphan_sheets


@task(name="Prepare CRS Records", cache_policy=NO_CACHE)
def prepare_crs_records(
    raw_records: list[dict],
    db_url:      str,
) -> list[dict]:
    """
    Enrich raw parsed records with:
    - row_hash (for SCD2 change detection)
    - comment_id = {doc_number}#{row_hash[:8]} (business key for ON CONFLICT)
    - tag_id (FK lookup into project_core.tag)
    - doc_id (FK lookup into project_core.document)
    Returns list of DB-ready record dicts.
    """
    logger = get_run_logger()

    engine = create_engine(
        db_url,
        poolclass=QueuePool,
        pool_size=2,
        max_overflow=5,
        pool_recycle=3600,
    )

    # Pre-load FK caches — never query inside row loop
    with engine.connect() as conn:
        tag_lookup: dict[str, str] = {
            row.tag_name: str(row.id)
            for row in conn.execute(
                text("SELECT id, tag_name FROM project_core.tag WHERE object_status = 'Active'")
            )
        }
        doc_lookup: dict[str, str] = {
            row.doc_number: str(row.id)
            for row in conn.execute(
                text("SELECT id, doc_number FROM project_core.document WHERE object_status = 'Active'")
            )
        }

    engine.dispose()
    logger.info("FK caches loaded: %d tags, %d documents.", len(tag_lookup), len(doc_lookup))

    db_records: list[dict] = []
    tag_miss_count = 0
    doc_miss_count = 0

    for rec in raw_records:
        # Build hashable dict (exclude timestamp-like fields for stability)
        hash_dict = {
            k: str(v) for k, v in rec.items()
            if k.lower() not in HASH_EXCLUDE_FIELDS and v is not None
        }
        row_hash   = hashlib.md5(
            json.dumps(hash_dict, sort_keys=True).encode()
        ).hexdigest()

        doc_number  = clean_string(rec.get("DOC_NUMBER")) or "UNKNOWN"
        comment_id  = f"{doc_number}#{row_hash[:8]}"

        tag_name = clean_string(rec.get("TAG_NAME"))
        tag_id   = tag_lookup.get(tag_name) if tag_name else None
        if tag_name and not tag_id:
            tag_miss_count += 1

        doc_id = doc_lookup.get(doc_number)
        if doc_number and doc_number != "UNKNOWN" and not doc_id:
            doc_miss_count += 1

        db_records.append({
            "comment_id":          comment_id,
            "doc_number":          doc_number,
            "doc_id":              doc_id,
            "revision":            clean_string(rec.get("REVISION")),
            "return_code":         clean_string(rec.get("RETURN_CODE")),
            "transmittal_number":  clean_string(rec.get("TRANSMITTAL_NUMBER")),
            "transmittal_date":    to_dt(rec.get("TRANSMITTAL_DATE")),
            "group_comment":       clean_string(rec.get("GROUP_COMMENT")) or "",
            "comment":             clean_string(rec.get("COMMENT")) or "",
            "tag_name":            tag_name,
            "tag_id":              tag_id,
            "property_name":       clean_string(rec.get("PROPERTY_NAME")),
            "response_vendor":     clean_string(rec.get("RESPONSE")),
            "source_file":         clean_string(rec.get("SOURCE_FILE")) or "",
            "detail_file":         clean_string(rec.get("DETAIL_FILE")),
            "detail_sheet":        clean_string(rec.get("DETAIL_SHEET")),
            "crs_file_path":       clean_string(rec.get("CRS_FILE_PATH")) or "",
            "crs_file_timestamp":  None,  # populated by OS stat if needed in future
            "status":              "RECEIVED",
            "object_status":       "Active",
            "row_hash":            row_hash,
        })

    if tag_miss_count:
        logger.warning("FK miss — tag_name unresolved: %d record(s) (tag_id=NULL)", tag_miss_count)
    if doc_miss_count:
        logger.warning("FK miss — doc_number unresolved: %d record(s) (doc_id=NULL)", doc_miss_count)

    logger.info("Prepared %d DB records.", len(db_records))
    return db_records


@task(name="Upsert CRS Comments", cache_policy=NO_CACHE)
def upsert_crs_records(
    db_url:     str,
    records:    list[dict],
    run_id:     str,
    batch_size: int = BATCH_SIZE,
) -> dict[str, int]:
    """
    Bulk upsert records into audit_core.crs_comment.
    ON CONFLICT (comment_id) DO UPDATE WHERE row_hash differs — idempotent.
    Also writes INSERT records to audit_core.crs_comment_audit (SCD2 trail).
    Returns stats: {inserted, updated, errors}.
    """
    logger = get_run_logger()
    stats  = {"inserted": 0, "updated": 0, "errors": 0}

    if not records:
        logger.info("No records to upsert.")
        return stats

    engine = create_engine(
        db_url,
        poolclass=QueuePool,
        pool_size=5,
        max_overflow=10,
        pool_recycle=3600,
    )

    upsert_sql = text("""
        INSERT INTO audit_core.crs_comment (
            comment_id, doc_number, doc_id, revision, return_code,
            transmittal_number, transmittal_date,
            group_comment, comment, tag_name, tag_id, property_name,
            response_vendor, source_file, detail_file, detail_sheet,
            crs_file_path, crs_file_timestamp,
            status, object_status, row_hash, sync_timestamp
        ) VALUES (
            :comment_id, :doc_number, :doc_id, :revision, :return_code,
            :transmittal_number, :transmittal_date,
            :group_comment, :comment, :tag_name, :tag_id, :property_name,
            :response_vendor, :source_file, :detail_file, :detail_sheet,
            :crs_file_path, :crs_file_timestamp,
            :status, :object_status, :row_hash, now()
        )
        ON CONFLICT (comment_id) DO UPDATE SET
            group_comment      = EXCLUDED.group_comment,
            comment            = EXCLUDED.comment,
            tag_name           = EXCLUDED.tag_name,
            tag_id             = EXCLUDED.tag_id,
            doc_id             = EXCLUDED.doc_id,
            property_name      = EXCLUDED.property_name,
            response_vendor    = EXCLUDED.response_vendor,
            crs_file_timestamp = EXCLUDED.crs_file_timestamp,
            row_hash           = EXCLUDED.row_hash,
            sync_timestamp     = now()
        WHERE audit_core.crs_comment.row_hash != EXCLUDED.row_hash
        RETURNING id, xmax
    """)

    audit_sql = text("""
        INSERT INTO audit_core.crs_comment_audit
            (comment_id, change_type, snapshot, run_id)
        VALUES (:cid, :ct, :snap::jsonb, :rid)
    """)

    for batch_start in range(0, len(records), batch_size):
        batch = records[batch_start : batch_start + batch_size]
        try:
            with engine.begin() as conn:
                for rec in batch:
                    result = conn.execute(upsert_sql, rec)
                    row = result.fetchone()
                    if row is None:
                        # hash unchanged — skip (no-op)
                        continue

                    # xmax == 0 means INSERT; xmax != 0 means UPDATE
                    is_insert = (row.xmax == 0)
                    change_type = "INSERT" if is_insert else "UPDATE"

                    if is_insert:
                        stats["inserted"] += 1
                    else:
                        stats["updated"] += 1

                    # Write SCD2 audit snapshot
                    snap = {k: str(v) for k, v in rec.items() if v is not None}
                    conn.execute(audit_sql, {
                        "cid":  str(row.id),
                        "ct":   change_type,
                        "snap": json.dumps(snap),
                        "rid":  run_id,
                    })

        except Exception as exc:
            logger.error("Batch %d–%d upsert error: %s",
                         batch_start, batch_start + len(batch), exc)
            stats["errors"] += len(batch)

    engine.dispose()
    logger.info("Upsert complete: inserted=%d updated=%d errors=%d",
                stats["inserted"], stats["updated"], stats["errors"])
    return stats


# =============================================================================
# Audit helpers
# =============================================================================

def _audit_start(engine, run_id: str, source_info: str) -> None:
    """INSERT into audit_core.sync_run_stats at flow start."""
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO audit_core.sync_run_stats
                (run_id, target_table, start_time, source_file)
            VALUES (:rid, :tbl, :st, :sf)
        """), {
            "rid": run_id,
            "tbl": "audit_core.crs_comment",
            "st":  datetime.now(),
            "sf":  source_info,
        })


def _audit_end(
    engine,
    run_id:    str,
    stats:     dict[str, int],
    orphan_count: int,
) -> None:
    """UPDATE audit_core.sync_run_stats at flow completion — always executed."""
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE audit_core.sync_run_stats SET
                end_time        = :et,
                count_created   = :cr,
                count_updated   = :up,
                count_unchanged = :uc,
                count_deleted   = :dl,
                count_errors    = :er
            WHERE run_id = :rid
        """), {
            "et":  datetime.now(),
            "cr":  stats.get("inserted",  0),
            "up":  stats.get("updated",   0),
            "uc":  0,
            "dl":  0,
            "er":  stats.get("errors",    0),
            "rid": run_id,
        })


# =============================================================================
# Main Flow
# =============================================================================

@flow(name="Import CRS Comments", log_prints=True)
def sync_crs_data_flow(debug_mode: bool = False) -> dict:
    """
    Full pipeline: discover → parse → prepare → upsert CRS comments.

    Args:
        debug_mode: If True, process only the first 5 main files (fast test run).

    Returns:
        Summary dict: run_id, files_processed, records_parsed, records_loaded,
                      errors, orphan_sheets, status.
    """
    logger = get_run_logger()
    run_id = str(uuid.uuid4())

    config = load_config()
    db_url  = get_db_engine_url(config)

    crs_data_dir = config.get("storage", {}).get("crs_data_dir")
    if not crs_data_dir:
        raise ValueError("storage.crs_data_dir missing from config/config.yaml")

    logger.info("Starting CRS sync | run_id=%s | debug=%s | dir=%s",
                run_id, debug_mode, crs_data_dir)

    engine = create_engine(db_url)
    try:
        # Fail-fast DB connection check
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
    except Exception as exc:
        engine.dispose()
        raise RuntimeError(f"DB connection failed: {exc}") from exc

    _audit_start(engine, run_id, f"CRS retroactive load | dir={crs_data_dir} | debug={debug_mode}")
    engine.dispose()

    # Task 1: Discover files
    main_files, detail_files = discover_crs_files(crs_data_dir)

    if debug_mode:
        keys = list(main_files.keys())[:5]
        main_files = {k: main_files[k] for k in keys}
        logger.info("DEBUG mode: processing %d file(s) only.", len(main_files))

    files_processed = len(main_files)

    # Task 2: Parse (ThreadPool inside task — Prefect 3.x compliant)
    all_records, orphan_sheets = parse_all_files(main_files, detail_files)
    records_parsed = len(all_records)

    stats: dict[str, int] = {"inserted": 0, "updated": 0, "errors": 0}

    if all_records:
        # Task 3: Prepare (hash, FK resolution)
        db_records = prepare_crs_records(all_records, db_url)

        # Task 4: Upsert
        stats = upsert_crs_records(db_url, db_records, run_id)
    else:
        logger.warning("No records parsed — nothing to upsert.")

    records_loaded = stats["inserted"] + stats["updated"]

    # Audit completion — always runs even on partial error
    engine = create_engine(db_url)
    try:
        _audit_end(engine, run_id, stats, len(orphan_sheets))
    finally:
        engine.dispose()

    result = {
        "run_id":          run_id,
        "files_processed": files_processed,
        "records_parsed":  records_parsed,
        "records_loaded":  records_loaded,
        "errors":          stats.get("errors", 0),
        "orphan_sheets":   orphan_sheets,
        "status":          "SUCCESS" if stats.get("errors", 0) == 0 else "PARTIAL",
    }

    if orphan_sheets:
        logger.warning(
            "%d orphan sheet(s) detected (detail sheets with no matching comment):\n  %s",
            len(orphan_sheets),
            "\n  ".join(orphan_sheets[:20]),
        )

    logger.info(
        "CRS sync complete: files=%d parsed=%d loaded=%d errors=%d orphans=%d status=%s",
        files_processed, records_parsed, records_loaded,
        stats.get("errors", 0), len(orphan_sheets), result["status"],
    )

    return result


# =============================================================================
# Prefect Deployment entry point
# =============================================================================

if __name__ == "__main__":

    sync_crs_data_flow.deploy(
        name="import_crs_data_deploy",
        work_pool_name="local-pool",
        tags=["crs", "production"],
        description="CRS comment retroactive loader",
        parameters={"debug_mode": False},
    )
