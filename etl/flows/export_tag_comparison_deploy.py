"""Standalone Prefect flow: Tag Comparison Report (two-sheet XLSX)."""

import sys
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from prefect import flow, task, get_run_logger
from prefect.cache_policies import NO_CACHE
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

# Setup import path so tasks.* resolves from etl/
current_dir = Path(__file__).resolve().parent
script_root = current_dir.parent
if str(script_root) not in sys.path:
    sys.path.append(str(script_root))

try:
    from tasks.common import load_config, get_db_engine_url
except ImportError as e:
    print(f"[SKIP] {Path(__file__).name}: Could not import task modules. Details: {e}")
    sys.exit(0)

# Module-level config — same pattern as other export flows
config = load_config()
DB_URL = get_db_engine_url(config)
_EXPORT_DIR = config.get("storage", {}).get("export_dir", ".")

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# All columns loaded from project_core.tag (comparison candidates + metadata)
_TAG_COLUMNS: list[str] = [
    "source_id",
    "tag_name",
    "tag_status",
    "description",
    "parent_tag_raw",
    "tag_class_raw",
    "area_code_raw",
    "process_unit_raw",
    "discipline_code_raw",
    "po_code_raw",
    "design_company_name_raw",
    "manufacturer_company_raw",
    "vendor_company_raw",
    "article_code_raw",
    "model_part_raw",
    "safety_critical_item",
    "safety_critical_item_reason_awarded",
    "production_critical_item",
    "safety_critical_item_group",
    "company_raw",
    "requisition_code_raw",
    "part_of",
    "serial_no",
    "install_date",
    "startup_date",
    "warranty_end_date",
    "price",
    "tech_id",
    "ip_grade",
    "ex_class",
    "mc_package_code",
    "equip_no",
    "alias",
    "from_tag_raw",
    "to_tag_raw",
    "plant_raw",
    "sync_status",
    "object_status",
]

# Columns excluded from field-level comparison (anchor, hashes, metadata)
_EXCLUDE_FROM_COMPARISON: frozenset[str] = frozenset(
    {"source_id", "id", "row_hash", "sync_timestamp", "run_id", "sync_status", "object_status"}
)

# Fields used for actual field-level diffing
_DATA_COLS: list[str] = [c for c in _TAG_COLUMNS if c not in _EXCLUDE_FROM_COMPARISON]

# Date columns that need normalisation to DD.MM.YYYY for consistent comparison
_DATE_COLS: frozenset[str] = frozenset({"install_date", "startup_date", "warranty_end_date", "purchase_date"})

# Maps abbreviated snapshot JSON keys → full project_core.tag column names.
# The ETL writes short keys to keep JSONB compact; comparison needs full names.
# Source of truth: _SNAPSHOT_KEYS set in import_tag_data_deploy.py.
SNAPSHOT_KEY_MAP: dict[str, str] = {
    "tn":           "tag_name",
    "t_stat":       "tag_status",
    "dsc":          "description",
    "cls_raw":      "tag_class_raw",
    "area_raw":     "area_code_raw",
    "unit_raw":     "process_unit_raw",
    "disc_raw":     "discipline_code_raw",
    "po_raw":       "po_code_raw",
    "dco_raw":      "design_company_name_raw",
    "mfr_raw":      "manufacturer_company_raw",
    "v_raw":        "vendor_company_raw",
    "art_raw":      "article_code_raw",
    "m_raw":        "model_part_raw",
    "sci":          "safety_critical_item",
    "sci_rea":      "safety_critical_item_reason_awarded",
    "pci":          "production_critical_item",
    "sece_grp":     "safety_critical_item_group",
    "sn":           "serial_no",
    "inst":         "install_date",
    "start":        "startup_date",
    "warn":         "warranty_end_date",
    "prc":          "price",
    "tid":          "tech_id",
    "ip_gr":        "ip_grade",
    "ex_cls":       "ex_class",
    "mc_pkg":       "mc_package_code",
    "eq":           "equip_no",
    "als":          "alias",
    "from_tag_raw": "from_tag_raw",
    "to_tag_raw":   "to_tag_raw",
    "plt_raw":      "plant_raw",
    "prnt_raw":     "parent_tag_raw",
    "purch_dt":    "purchase_date",
    "co_raw":      "company_raw",
    "req_raw":     "requisition_code_raw",
    "part_of_raw": "part_of",
}

# ---------------------------------------------------------------------------
# SQL queries
# ---------------------------------------------------------------------------

_SQL_SNAPSHOT_FOR_DATE = """
/*
Purpose: Most recent per-tag snapshot on or before target_date.
Gate:    sync_timestamp::date <= target_date.
Note:    DISTINCT ON picks the latest history record per tag.
         Used for BOTH current and baseline states.
Changes: 2026-04-22 — Renamed from _SQL_BASELINE, now used for both sides.
*/
/*
Purpose: Most recent per-tag snapshot on or before baseline_date.
Gate:    sync_timestamp::date <= baseline_date.
Note:    DISTINCT ON picks the latest history record per tag.
Changes: 2026-04-16 — Initial implementation.
*/
SELECT DISTINCT ON (source_id)
    source_id,
    tag_name,
    row_hash,
    snapshot,
    sync_timestamp
FROM audit_core.tag_status_history
WHERE sync_timestamp::date <= :baseline_date
ORDER BY source_id, sync_timestamp DESC
"""

_SQL_MIN_HISTORY_DATE = """
SELECT MIN(sync_timestamp)::date AS min_date
FROM audit_core.tag_status_history
"""

_SQL_MAX_CURRENT_DATE = """
SELECT MAX(sync_timestamp)::date AS max_date
FROM audit_core.tag_status_history
"""

_SQL_NEAREST_BASELINE_DATE = """
SELECT MAX(sync_timestamp)::date AS nearest_date
FROM audit_core.tag_status_history
WHERE sync_timestamp::date <= :target_date
"""

# ---------------------------------------------------------------------------
# Style constants (openpyxl)
# ---------------------------------------------------------------------------

_FILL_GREEN_HEADER = PatternFill(fill_type="solid", fgColor="C6EFCE")   # _new headers
_FILL_BLUE_HEADER = PatternFill(fill_type="solid", fgColor="BDD7EE")    # _old headers
_FILL_GREY = PatternFill(fill_type="solid", fgColor="D9D9D9")           # neutral headers
_FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")         # Modified cell
_FILL_ROW_GREEN = PatternFill(fill_type="solid", fgColor="E2EFDA")      # Created row
_FILL_ROW_RED = PatternFill(fill_type="solid", fgColor="FCE4D6")        # Deleted row

_FONT_GREEN = Font(name="Calibri", size=11, bold=True, color="276221")
_FONT_BLUE = Font(name="Calibri", size=11, bold=True, color="1F497D")
_FONT_BLACK_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
_FONT_DEFAULT = Font(name="Calibri", size=11)

_CENTER = Alignment(horizontal="center")

# Sort priority for Comparison_Result
_RESULT_SORT_ORDER: dict[str, int] = {
    "Modified": 0,
    "Created": 1,
    "Deleted": 2,
    "No Changes": 3,
}

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def _to_str(value) -> str:
    """Convert any value to string; None/NaN/NaT map to empty string."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def _escape_formula(value: str) -> str:
    """Prepend a space if value starts with '=' to prevent formula injection."""
    if value.startswith("="):
        return " " + value
    return value


def _fmt_date(v) -> str:
    """Normalise any date string to DD.MM.YYYY; return '' if unparseable."""
    if not v:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    from datetime import datetime as _dt
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return _dt.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return s


def _safe_cell(value) -> str:
    """Convert value to safe cell string (None→'', formula-escaped)."""
    return _escape_formula(_to_str(value))


def _normalise_date_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise date columns to DD.MM.YYYY string for consistent comparison."""
    for col in _DATE_COLS:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


def _autofit_columns(
    ws, min_width: int = 12, max_width: int = 50, padding: int = 2
) -> None:
    """Set each column width based on the longest cell value in that column."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(
            max_width, max(min_width, max_len + padding)
        )


def _resolve_max_current_date(engine: Engine) -> date:
    with engine.connect() as conn:
        row = conn.execute(text(_SQL_MAX_CURRENT_DATE)).fetchone()
    if row is None or row[0] is None:
        raise ValueError("project_core.tag contains no active records.")
    return row[0]


def _numeric_sort_key(val: str) -> tuple:
    """Sort key that orders numeric strings numerically, non-numeric strings last."""
    try:
        return (0, int(val))
    except (ValueError, TypeError):
        return (1, val)


# ---------------------------------------------------------------------------
# Comparison logic
# ---------------------------------------------------------------------------


def _get_comparison_result(row: pd.Series, data_cols: list[str]) -> str:
    """
    Classify a merged row as Modified, Created, Deleted, or No Changes.

    Mirrors ReportComparisionTool.py logic exactly:
    - Fast path: identical row_hash → No Changes (avoids field scan)
    - VOID tag_status → Deleted
    - left_only (not in baseline) → Created
    - right_only (not in current) → Deleted

    Args:
        row: Series from the outer-merged DataFrame.
        data_cols: Column names used for field-level comparison.

    Returns:
        One of 'Modified', 'Created', 'Deleted', 'No Changes', 'Unknown'.
    """
    merge_flag = row["_merge"]
    tag_status_new = _to_str(row.get("tag_status_new", ""))

    if merge_flag == "both":
        # Fast path: skip field scan when hash is identical
        if row.get("row_hash_new") == row.get("row_hash_old"):
            return "No Changes"
        if tag_status_new == "VOID":
            return "Deleted"
        changed = any(
            _to_str(row.get(f"{c}_new", "")) != _to_str(row.get(f"{c}_old", ""))
            for c in data_cols
        )
        return "Modified" if changed else "No Changes"

    if merge_flag == "left_only":
        return "Deleted" if tag_status_new == "VOID" else "Created"

    if merge_flag == "right_only":
        return "Deleted"

    return "Unknown"


# ---------------------------------------------------------------------------
# Prefect tasks
# ---------------------------------------------------------------------------


@task(name="load-snapshot-for-date", retries=1, cache_policy=NO_CACHE)
def load_snapshot_for_date(engine: Engine, target_date: date) -> pd.DataFrame:
    """
    Load the most recent snapshot per tag recorded on or before target_date.

    Uses identical logic for both current and baseline states.

    Args:
        engine: SQLAlchemy engine connected to engineering_core.
        target_date: Upper bound (inclusive) for sync_timestamp.

    Returns:
        DataFrame keyed on source_id with row_hash and all data_cols.

    Raises:
        ValueError: If no history records exist on or before target_date.
    """
    logger = get_run_logger()
    with engine.connect() as conn:
        df = pd.read_sql(
            text(_SQL_SNAPSHOT_FOR_DATE).bindparams(target_date=str(target_date)),
            conn,
        )

    if df.empty:
        with engine.connect() as conn:
            row = conn.execute(text(_SQL_MIN_HISTORY_DATE)).fetchone()
        min_ts = row[0] if row else "unknown"
        raise ValueError(
            f"No history records found on or before {target_date}. "
            f"Earliest available: {min_ts}"
        )

    # Expand snapshot JSONB → one column per key
    snapshot_expanded: pd.DataFrame = df["snapshot"].apply(
        lambda s: s if isinstance(s, dict) else {}
    ).apply(pd.Series)

    # Debug: log raw snapshot keys
    raw_keys = set(snapshot_expanded.columns)
    logger.debug(f"Raw snapshot keys found: {sorted(raw_keys)}")

    # Rename abbreviated keys → full _DATA_COLS names
    snapshot_expanded = snapshot_expanded.rename(
        columns={k: v for k, v in SNAPSHOT_KEY_MAP.items() if k in snapshot_expanded.columns}
    )

    # Debug: log which _DATA_COLS are missing after rename
    missing_cols = [c for c in _DATA_COLS if c not in snapshot_expanded.columns]
    if missing_cols:
        logger.debug(
            f"Snapshot fields missing after rename (will be filled with ''): {missing_cols}"
        )

    # Ensure all data_cols are present; fill any missing ones with ''
    for col in _DATA_COLS:
        if col not in snapshot_expanded.columns:
            snapshot_expanded[col] = ""

    snapshot_expanded = snapshot_expanded.reindex(columns=_DATA_COLS, fill_value="")

    # Normalise date columns from JSON strings
    snapshot_expanded = _normalise_date_cols(snapshot_expanded)

    # Replace sentinel strings with empty string
    _SENTINEL_VALUES = {'None', 'nan', 'NaT', 'null', 'NULL'}
    for col in snapshot_expanded.columns:
        snapshot_expanded[col] = snapshot_expanded[col].apply(
            lambda v: "" if _to_str(v).strip() in _SENTINEL_VALUES else _to_str(v).strip()
        )

    # Combine anchor columns with expanded snapshot
    result_df = pd.concat(
        [
            df[["source_id", "row_hash"]].reset_index(drop=True),
            snapshot_expanded.reset_index(drop=True),
        ],
        axis=1,
    )
    logger.info(f"Loaded {len(result_df)} snapshots as of {target_date}")
    return result_df


# ---------------------------------------------------------------------------
# XLSX sheet builders
# ---------------------------------------------------------------------------


def _build_full_comparison_sheet(
    ws, merged: pd.DataFrame, data_cols: list[str]
) -> None:
    """
    Write Sheet 1 "Full Comparison".

    Column layout: [_new block] | source_id | [_old block] | Comparison_Result
    Colour rules:
      - _new header: light green / dark green bold
      - _old header: light blue / dark blue bold
      - source_id / Comparison_Result header: grey / black bold
      - Modified cell: yellow on cells where field_new != field_old
      - Created row: light green on all _new cells
      - Deleted row: light red on all _old cells
    Sort order: Modified → Created → Deleted → No Changes, then source_id ASC.

    Args:
        ws: openpyxl Worksheet to write into.
        merged: Outer-merged DataFrame with Comparison_Result column.
        data_cols: Ordered list of field names used for comparison.
    """
    new_headers = [f"{c}_new" for c in data_cols]
    old_headers = [f"{c}_old" for c in data_cols]
    all_headers = new_headers + ["source_id"] + old_headers + ["Comparison_Result"]

    # --- Header row ---
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = _CENTER
        if header.endswith("_new"):
            cell.fill = _FILL_GREEN_HEADER
            cell.font = _FONT_GREEN
        elif header.endswith("_old"):
            cell.fill = _FILL_BLUE_HEADER
            cell.font = _FONT_BLUE
        else:
            cell.fill = _FILL_GREY
            cell.font = _FONT_BLACK_BOLD

    # --- Sort ---
    df_sorted = merged.copy()
    df_sorted["_sort_key"] = df_sorted["Comparison_Result"].map(
        lambda r: _RESULT_SORT_ORDER.get(r, 99)
    )
    df_sorted["_source_id_int"] = pd.to_numeric(
        df_sorted["source_id"], errors="coerce"
    )
    df_sorted.sort_values(
        ["_sort_key", "_source_id_int"], na_position="last", inplace=True
    )
    df_sorted.drop(columns=["_source_id_int"], inplace=True)
    df_sorted.reset_index(drop=True, inplace=True)

    # --- Data rows ---
    for excel_row, (_, row) in enumerate(df_sorted.iterrows(), start=2):
        result = _to_str(row.get("Comparison_Result", ""))

        for col_idx, header in enumerate(all_headers, start=1):
            cell_val = _safe_cell(row.get(header, ""))
            cell = ws.cell(row=excel_row, column=col_idx, value=cell_val)
            cell.font = _FONT_DEFAULT

            # Cell-level fill
            if result == "Modified":
                # Highlight only cells where the field actually changed
                if header.endswith("_new") or header.endswith("_old"):
                    base = header[:-4]  # strip '_new' or '_old'
                    val_new = _to_str(row.get(f"{base}_new", ""))
                    val_old = _to_str(row.get(f"{base}_old", ""))
                    if val_new != val_old:
                        cell.fill = _FILL_YELLOW
            elif result == "Created" and header.endswith("_new"):
                cell.fill = _FILL_ROW_GREEN
            elif result == "Deleted" and header.endswith("_old"):
                cell.fill = _FILL_ROW_RED

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws)


def _build_changes_only_sheet(
    ws, merged: pd.DataFrame, data_cols: list[str]
) -> None:
    """
    Write Sheet 2 "Changes Only".

    Only fields that actually changed are shown per Modified/Created/Deleted tag:
    - Modified: rows where val_old != val_new
    - Created: rows where val_new != ''
    - Deleted: rows where val_old != ''

    Columns: TAG_NAME | EIS_FIELD | VALUE_OLD | VALUE_NEW | STATUS | _TYPE (hidden)
    _TYPE (col F) holds the comparison type (Modified/Created/Deleted) and drives
    conditional formatting rules so colours survive Excel filtering.

    Args:
        ws: openpyxl Worksheet to write into.
        merged: Outer-merged DataFrame with Comparison_Result column.
        data_cols: Ordered list of field names used for comparison.
    """
    headers = ["TAG_NAME", "EIS_FIELD", "VALUE_OLD", "VALUE_NEW", "STATUS", "_TYPE"]
    header_fills = {
        "TAG_NAME":  _FILL_GREY,
        "EIS_FIELD": _FILL_GREY,
        "VALUE_OLD": _FILL_BLUE_HEADER,
        "VALUE_NEW": _FILL_GREEN_HEADER,
        "STATUS":    _FILL_GREY,
        "_TYPE":     _FILL_GREY,
    }
    header_fonts = {
        "TAG_NAME":  _FONT_BLACK_BOLD,
        "EIS_FIELD": _FONT_BLACK_BOLD,
        "VALUE_OLD": _FONT_BLUE,
        "VALUE_NEW": _FONT_GREEN,
        "STATUS":    _FONT_BLACK_BOLD,
        "_TYPE":     _FONT_BLACK_BOLD,
    }

    # --- Header row ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fills[header]
        cell.font = header_fonts[header]
        cell.alignment = _CENTER

    changed = merged[
        merged["Comparison_Result"].isin(["Modified", "Created", "Deleted"])
    ]

    excel_row = 2

    for _, row in changed.iterrows():
        comparison = _to_str(row.get("Comparison_Result", ""))
        tag_name = _to_str(row.get("tag_name_new", "")) or _to_str(
            row.get("tag_name_old", "")
        )

        # Collect only truly changed fields (alphabetical order)
        tag_field_rows: list[tuple[str, str, str]] = []
        for field in sorted(data_cols):
            if comparison == "Created":
                val_old = ""
                val_new = _to_str(row.get(f"{field}_new", ""))
                if val_new == "":
                    continue
            elif comparison == "Deleted":
                val_old = _to_str(row.get(f"{field}_old", ""))
                val_new = ""
                if val_old == "":
                    continue
            else:  # Modified
                val_old = _to_str(row.get(f"{field}_old", ""))
                val_new = _to_str(row.get(f"{field}_new", ""))
                if val_old == val_new:
                    continue

            tag_field_rows.append((field, val_old, val_new))

        for field, val_old, val_new in tag_field_rows:
            row_values = [
                tag_name,
                field.upper(),
                _escape_formula(val_old),
                _escape_formula(val_new),
                "✓ CHANGED",
                comparison,
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                cell.font = _FONT_DEFAULT
            excel_row += 1

        # Blank separator row between tag groups
        excel_row += 1

    # Apply FormulaRule-based conditional formatting so colours survive filtering
    if excel_row > 2:
        max_row = excel_row - 1
        ws.conditional_formatting.add(
            f"C2:D{max_row}",
            FormulaRule(formula=['=$F2="Modified"'], fill=_FILL_YELLOW),
        )
        ws.conditional_formatting.add(
            f"D2:D{max_row}",
            FormulaRule(formula=['=$F2="Created"'], fill=_FILL_ROW_GREEN),
        )
        ws.conditional_formatting.add(
            f"C2:C{max_row}",
            FormulaRule(formula=['=$F2="Deleted"'], fill=_FILL_ROW_RED),
        )

    # Hide the helper column used by FormulaRule
    ws.column_dimensions["F"].hidden = True

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(5)}1"
    _autofit_columns(ws, min_width=15, max_width=60, padding=2)


# ---------------------------------------------------------------------------
# Prefect flow
# ---------------------------------------------------------------------------


@flow(name="export_tag_comparison", log_prints=True)
def export_tag_comparison_flow(
    current_date: Optional[date] = None,
    baseline_date: Optional[date] = None,
    doc_revision: str = "A01",
    output_dir: Optional[str] = None,
) -> str:
    """
    Generates a two-sheet XLSX tag comparison report.

    Args:
        current_date:  Reference date for "new" state.
                       Default: resolved from DB as MAX(sync_timestamp)::date
                       from project_core.tag WHERE object_status = 'Active'.
        baseline_date: Reference date for "old" (baseline) state.
                       Default: nearest available history date on or before
                       (current_date minus 1 calendar month) — queried from
                       audit_core.tag_status_history. Raises ValueError if no
                       history records exist before that arithmetic target.
        doc_revision:  Document revision code embedded in the output filename
                       (e.g. 'A01', 'B02'). Default: 'A01'.
        output_dir:    Override output directory.
                       Default: same output root as other export flows.

    Returns:
        Absolute path of the written XLSX file.

    Raises:
        ValueError: If baseline_date >= current_date (comparison of a snapshot
                    with itself produces no meaningful output).
        ValueError: If no active tags exist in project_core.tag as of
                    current_date (would produce misleading all-Deleted output).
        ValueError: If current_date cannot be resolved from DB (no active tags).
        ValueError: If baseline_date cannot be resolved from DB (no history
                    records before the arithmetic target date).

    Output filename:
        Tag&Equipment-register_compare_{doc_revision}.xlsx
        Example: Tag&Equipment-register_compare_A37.xlsx

    Example:
        >>> export_tag_comparison_flow(
        ...     current_date=date(2026, 4, 16),
        ...     baseline_date=date(2026, 3, 16),
        ...     doc_revision="A02",
        ... )
        '/mnt/shared-data/.../Tag&Equipment-register_compare_A02.xlsx'
    """
    logger = get_run_logger()

    # Engine must be created before date resolution (DB queries needed for defaults)
    engine = create_engine(DB_URL, pool_pre_ping=True)

    # Resolve current_date default from DB MAX
    _current_date_explicit = current_date is not None
    if current_date is None:
        current_date = _resolve_max_current_date(engine)

    # Resolve baseline_date default using DB proximity query
    _baseline_date_explicit = baseline_date is not None
    if baseline_date is None:
        _target = current_date - relativedelta(months=1)
        with engine.connect() as conn:
            row = conn.execute(
                text(_SQL_NEAREST_BASELINE_DATE).bindparams(target_date=str(_target))
            ).fetchone()
        if row is None or row[0] is None:
            raise ValueError(
                f"No tag history found on or before {_target} "
                f"(1 month before current_date={current_date}). "
                f"Cannot resolve default baseline_date."
            )
        baseline_date = row[0]
        if baseline_date != _target:
            logger.warning(
                f"No history on exact target {_target}; "
                f"using nearest available date: {baseline_date}"
            )

    # Guard: baseline must be strictly before current
    if baseline_date >= current_date:
        raise ValueError(
            f"baseline_date ({baseline_date}) must be strictly before "
            f"current_date ({current_date}). "
            f"Comparison of a snapshot with itself produces no meaningful output."
        )

    current_date_str = current_date.strftime("%Y-%m-%d")
    filename = f"Tag&Equipment-register_compare_{doc_revision}.xlsx"
    output_path = Path(output_dir or _EXPORT_DIR) / filename

    # Load both states
    # Load both states from snapshots
    df_new = load_snapshot_for_date(engine, current_date)
    df_old = load_snapshot_for_date(engine, baseline_date)

    # Sanity: confirm baseline snapshot keys were resolved correctly
    if not df_old.empty:
        sample = df_old[["tag_name", "tag_class_raw", "area_code_raw"]].head(3)
        logger.info(f"Baseline snapshot sample (first 3 rows):\n{sample.to_string(index=False)}")

    logger.info(
        f"Current date: {current_date} "
        f"({'explicit' if _current_date_explicit else 'resolved from DB MAX'})"
        f" | Baseline date: {baseline_date} "
        f"({'explicit' if _baseline_date_explicit else 'resolved from DB nearest'})"
    )
    logger.info(
        f"Tags in current: {len(df_new)} | Tags in baseline: {len(df_old)}"
    )

    # Trim to join key + row_hash + _DATA_COLS only
    new_keep = ["source_id", "row_hash"] + [
        c for c in _DATA_COLS if c in df_new.columns
    ]
    old_keep = ["source_id", "row_hash"] + [
        c for c in _DATA_COLS if c in df_old.columns
    ]
    df_new_trim = df_new[new_keep].copy()
    df_old_trim = df_old[old_keep].copy()

    # FIXED: Outer merge on source_id (stable spreadsheet identifier, not DB auto-increment id)
    # source_id is immutable from the source spreadsheet system and is the correct join key.
    # The DB auto-increment 'id' column is NOT used for comparison.
    logger.info("Merging on source_id (stable spreadsheet identifier, not DB auto-increment id)")
    merged = pd.merge(
        df_new_trim,
        df_old_trim,
        on="source_id",
        suffixes=("_new", "_old"),
        how="outer",
        indicator=True,
    )
    merged["source_id"] = merged["source_id"].fillna("").astype(str)

    # Classify every row
    merged["Comparison_Result"] = merged.apply(
        lambda row: _get_comparison_result(row, _DATA_COLS),
        axis=1,
    )

    n_mod = int((merged["Comparison_Result"] == "Modified").sum())
    n_cre = int((merged["Comparison_Result"] == "Created").sum())
    n_del = int((merged["Comparison_Result"] == "Deleted").sum())
    n_nc = int((merged["Comparison_Result"] == "No Changes").sum())

    logger.info(
        f"Modified: {n_mod} | Created: {n_cre} | Deleted: {n_del} | No Changes: {n_nc}"
    )

    # Build workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Full Comparison"
    _build_full_comparison_sheet(ws1, merged, _DATA_COLS)

    ws2 = wb.create_sheet(title="Changes Only")
    _build_changes_only_sheet(ws2, merged, _DATA_COLS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    logger.info(f"Output: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    _REPO_ROOT = Path(__file__).resolve().parent.parent.parent
    export_tag_comparison_flow.from_source(
        source=str(_REPO_ROOT),
        entrypoint="etl/flows/export_tag_comparison_deploy.py:export_tag_comparison_flow",
    ).deploy(
        name="export_tag_comparison_deploy",
        work_pool_name="default-agent-pool",
    )
